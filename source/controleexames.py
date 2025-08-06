import datetime
from pathlib import Path
from openpyxl import load_workbook
import json
import shutil
import re

class ControleExames:
    def __init__(self, pasta_export="exportados"):
        self.raiz = Path("relacoes")
        self.raiz_modelos = Path(r"relacoes/modelos")
        self.modelo = Path(r"relacoes/modelos/relacao.xlsx")
        self.json = Path("Relacao_empresas.json")
        self.pasta_export = Path(pasta_export)
        if not self.raiz.exists():
            self.raiz.mkdir(exist_ok=True)
        if not self.raiz_modelos.exists():
            self.raiz_modelos.mkdir(exist_ok=True)
        if not self.pasta_export.exists():
            self.pasta_export.mkdir(exist_ok=True)
        if datetime.datetime.now().day == 1:
            self.fechar_mes()
        self.init_json()

    def limpar_nome(self, nome):
        """Remove caracteres inválidos para nomes de arquivo no Windows"""
        return re.sub(r'[\\/*?:"<>|]', "", nome)

    def init_json(self):   
        dados = {} 
        if not self.json.exists():
            with open("Relacao_empresas.json", "w", encoding="utf-8") as f:
                json.dump(dados, f, indent=4, ensure_ascii=False)

    def _carregar_json(self, empresa):
        try:
            caminho = Path("Relacao_empresas.json")
            with open(caminho, "r", encoding="utf-8") as f:
                dados = json.load(f)
                return dados[empresa]
        except:
            self._salvar_json(empresa, 6)
            with open(caminho, "r", encoding="utf-8") as f:
                dados = json.load(f)
                return dados[empresa]

    def _salvar_json(self, empresa, value):
        with open("Relacao_empresas.json", "r", encoding="utf-8") as f:
            dados = json.load(f)
        dados[empresa] = value
        with open("Relacao_empresas.json", "w", encoding="utf-8") as f:
            json.dump(dados, f, ensure_ascii=False, indent=4)

    def _criar_planilha(self, empresa: str, cnpj: str, exames: list[str], data_exame: str):
        try:
            print(f"Iniciando criação da planilha para: {empresa}")
            empresa_limpa = self.limpar_nome(empresa)
            exames_limpos = self.limpar_nome(", ".join(exames))
            data_limpa = data_exame.replace("/", "-")
            if not empresa_limpa.strip() or empresa_limpa.strip().isdigit():
                print("Nome da empresa inválido, não salvando arquivo.")
                return

            wb = load_workbook(self.modelo)
            ws = wb.active
            ws["B2"] = empresa
            ws["B3"] = cnpj

            empresa_arquivo = f"{empresa_limpa} - {exames_limpos} - {data_limpa}.xlsx"
            saida = self.raiz
            output_path = saida / empresa_arquivo
            output_path.parent.mkdir(parents=True, exist_ok=True)
            print(f"Salvando planilha em: {output_path.absolute()}")
            wb.save(output_path)
            print(f"Planilha criada com sucesso!")
        except Exception as e:
            print(f"ERRO ao criar planilha: {e}")
            raise e

    def registrar_exames(self, empresa: str, cnpj: str, nome: str, exames: list[str], data_exame: str):
        try:
            print(f"Registrando exames para: {empresa} - {nome}")
            empresa_limpa = self.limpar_nome(empresa)
            exames_limpos = self.limpar_nome(", ".join(exames))
            data_limpa = data_exame.replace("/", "-")
            if not empresa_limpa.strip() or empresa_limpa.strip().isdigit():
                print("Nome da empresa inválido, não salvando arquivo.")
                return

            nome_arquivo = f"{empresa_limpa} - {exames_limpos} - {data_limpa}.xlsx"
            relacao = self.raiz / nome_arquivo
            print(f"Procurando arquivo: {relacao.absolute()}")

            if not relacao.exists():
                print("Arquivo não existe, criando...")
                self._criar_planilha(empresa, cnpj, exames, data_exame)
                self._salvar_json(empresa, 6)

            wb = load_workbook(relacao)
            ws = wb.active
            inicial_value = self._carregar_json(empresa)

            ws[f"B{inicial_value}"] = nome
            ws[f"D{inicial_value}"] = ", ".join(exames)
            ws[f"K{inicial_value}"] = data_exame

            self._salvar_json(empresa, inicial_value + 1)
            wb.save(relacao)
            print(f"Exames registrados com sucesso em: {relacao.absolute()}")
        except Exception as e:
            print(f"ERRO ao registrar exames: {e}")
            raise e

    def fechar_mes(self):
        data_atual = datetime.datetime.now()
        if data_atual.day != 1:
            return  # só executa se for dia 1 do mês

        if not self.json.exists():
            return

        with open(self.json, "r", encoding="utf-8") as f:
            empresas_json = json.load(f)

        for empresa in empresas_json:
            empresa_limpa = self.limpar_nome(empresa)
            arquivos = list(self.raiz.glob(f"{empresa_limpa} - *_{data_atual.strftime('%m-%Y')}.xlsx"))

            if arquivos:
                caminho_arquivo_atual = arquivos[0]
                destino = self.pasta_export / caminho_arquivo_atual.name
                shutil.move(str(caminho_arquivo_atual), str(destino))
                cnpj = caminho_arquivo_atual.name.split(" - ")[1].split("_")[0]
            else:
                cnpj = ""
            self._criar_planilha(empresa, cnpj, ["EXAMES"], data_atual.strftime("%d-%m-%Y"))
            self._salvar_json(empresa, 5)