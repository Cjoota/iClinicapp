import flet as ft
from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook
from Interfaces.sidebar import Sidebar
from Interfaces.telaresize import Responsive

def listar_exames_por_empresa_controle():
    relacoes_dir = Path("relacoes")
    exames_por_empresa = defaultdict(list)

    if not relacoes_dir.exists():
        return {}

    for arquivo in relacoes_dir.glob("*.xlsx"):
        if "modelo" in arquivo.name.lower():
            continue  # Pula arquivos de modelo

        empresa_nome = arquivo.stem.rsplit("_", 1)[0].replace("-", " ")
        try:
            wb = load_workbook(arquivo)
            ws = wb.active
            # Supondo que os dados começam na linha 6
            for row in ws.iter_rows(min_row=6, values_only=True):
                if not row or len(row) < 11:
                    continue
                nome_colaborador = row[1]  # Coluna B
                exames = row[3]            # Coluna D
                data_exame = row[10]       # Coluna K
                if nome_colaborador and exames and data_exame:
                    exames_por_empresa[empresa_nome].append({
                        "exame": exames,
                        "colaborador": nome_colaborador,
                        "data": str(data_exame),
                        "hora": "",  # Se quiser adicionar hora, ajuste aqui
                        "arquivo": arquivo
                    })
        except Exception as e:
            print(f"Erro ao ler arquivo {arquivo}: {e}")
            continue

    return exames_por_empresa

class Relacoes:
    def __init__(self, page: ft.Page):
        self.page = page
        self.responsive = Responsive(self.page)
        self.page.title = "Relações"
        self.sidebar = Sidebar(self.page)
        self.expansion_list = ft.ExpansionPanelList()
        self.empresas_exames = listar_exames_por_empresa_controle()
        self.search_field = ft.TextField(
            label="Pesquisar empresa...",
            on_change=self.filtrar_empresas,
            width=400,
            prefix_icon=ft.Icons.SEARCH
        )
        self.carregar_empresas()

    def carregar_empresas(self, filtro=""):
        self.expansion_list.controls.clear()
        for empresa, exames in self.empresas_exames.items():
            if filtro.lower() in empresa.lower():
                self.expansion_list.controls.append(self.criar_painel(empresa, exames))
        self.page.update()

    def filtrar_empresas(self, e):
        filtro = e.control.value
        self.carregar_empresas(filtro)

    def excluir_exame(self, arquivo, colaborador, exame_texto, data):
        try:
            wb = load_workbook(arquivo)
            ws = wb.active
            for row_num in range(6, ws.max_row + 1):
                if (ws[f"B{row_num}"].value == colaborador and
                    ws[f"D{row_num}"].value == exame_texto and
                    str(ws[f"K{row_num}"].value) == str(data)):
                    ws[f"B{row_num}"] = None
                    ws[f"D{row_num}"] = None
                    ws[f"K{row_num}"] = None
                    break
            wb.save(arquivo)
            print(f"Exame excluído com sucesso")
        except Exception as e:
            print(f"Erro ao excluir exame: {e}")

        self.empresas_exames = listar_exames_por_empresa_controle()
        self.carregar_empresas(self.search_field.value if hasattr(self, 'search_field') else "")

    def criar_painel(self, empresa, exames):
        exames_widgets = []
        if exames:
            exames_widgets.append(
                ft.Text(
                    "Exames realizados:",
                    weight=ft.FontWeight.BOLD,
                    size=16,
                    color=ft.Colors.BLUE_700,
                    text_align=ft.TextAlign.LEFT
                )
            )
            for exame in exames:
                exames_widgets.append(
                    ft.Row(
                        [
                            ft.Text(
                                f"Colaborador: {exame['colaborador']} | Exame: {exame['exame']} | Data: {exame['data']} {exame['hora']}",
                                size=14,
                                color=ft.Colors.GREY_900,
                                text_align=ft.TextAlign.LEFT,
                                expand=True
                            ),
                            ft.IconButton(
                                icon=ft.Icons.DELETE,
                                icon_color=ft.Colors.RED,
                                tooltip="Excluir exame",
                                on_click=lambda e, arquivo=exame['arquivo'], colaborador=exame['colaborador'],
                                       exame_texto=exame['exame'], data=exame['data']:
                                       self.excluir_exame(arquivo, colaborador, exame_texto, data)
                            )
                        ],
                        alignment=ft.MainAxisAlignment.SPACE_BETWEEN
                    )
                )
        else:
            exames_widgets.append(
                ft.Text(
                    "Nenhum exame realizado.",
                    italic=True,
                    size=14,
                    color=ft.Colors.GREY_500,
                    text_align=ft.TextAlign.LEFT
                )
            )

        return ft.ExpansionPanel(
            header=ft.Text(empresa, size=18, weight=ft.FontWeight.BOLD, color=ft.Colors.GREEN_800),
            content=ft.Container(
                content=ft.Column(
                    exames_widgets,
                    spacing=6,
                    alignment=ft.MainAxisAlignment.START,
                    horizontal_alignment=ft.CrossAxisAlignment.START
                ),
                bgcolor=ft.Colors.WHITE,
                border_radius=12,
                border=ft.border.all(1, ft.Colors.GREY_200),
                margin=ft.margin.symmetric(vertical=8, horizontal=0),
                padding=ft.padding.all(16)
            )
        )

    def build_view(self):
        conteudo = ft.Column(
            [
                ft.Text("Relações de Empresas", size=28, weight=ft.FontWeight.BOLD, color=ft.Colors.GREEN_900),
                ft.Divider(height=2, color=ft.Colors.GREY_300),
                self.search_field,
                self.expansion_list
            ],
            expand=True,
            scroll=ft.ScrollMode.AUTO,
            spacing=20
        )

        if self.responsive.is_mobile():
            return ft.Column(
                [
                    ft.Row(
                        [self.sidebar.build()],
                        alignment=ft.MainAxisAlignment.CENTER,
                        vertical_alignment=ft.CrossAxisAlignment.CENTER,
                        width=self.responsive.content_width()
                    ),
                    ft.Column(
                        [ft.Container(content=conteudo, padding=0)],
                        alignment=ft.MainAxisAlignment.CENTER,
                        horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                        spacing=0
                    )
                ],
                alignment=ft.MainAxisAlignment.CENTER,
                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                spacing=0
            )
        elif self.responsive.is_tablet():
            return ft.Column(
                [
                    ft.Row(
                        [self.sidebar.build()],
                        alignment=ft.MainAxisAlignment.CENTER,
                        vertical_alignment=ft.CrossAxisAlignment.CENTER
                    ),
                    ft.Column(
                        [conteudo],
                        alignment=ft.MainAxisAlignment.CENTER,
                        horizontal_alignment=ft.CrossAxisAlignment.CENTER
                    )
                ],
                alignment=ft.MainAxisAlignment.CENTER,
                horizontal_alignment=ft.CrossAxisAlignment.CENTER
            )
        else:
            return ft.Row(
                [
                    ft.Column(
                        [self.sidebar.build()],
                        alignment=ft.MainAxisAlignment.START
                    ),
                    ft.Container(content=conteudo, expand=True)
                ],
                width=self.page.width,
                height=self.page.height,
                alignment=ft.MainAxisAlignment.START
            )