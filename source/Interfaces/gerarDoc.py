from openpyxl import load_workbook
from funcoes import puxardados, verempresa
import flet as ft
import datetime as dt
from pathlib import Path
from Interfaces.telaresize import Responsive
from Interfaces.main_interface import Main_interface
from Interfaces.sidebar import Sidebar
from datetime import date
class Gerardoc:
        def __init__(self,page:ft.Page) -> None:
            self.page = page
            self.responsive = Responsive(self.page)
            self.sidebar = Sidebar(self.page)
            self.main = Main_interface(self.page)
            self.page.clean()
            self.modelos_excel = self.carregar_modelos_excel()
            self.dataselect = None
            self.empresas_drop = verempresa()
            self.risk = []
            self.risk_quimico = []
            self.risk_fisico = []
            self.risk_biologico = []
            self.risk_ergonomico = []
            self.page.on_resized = self.on_resize

        def limitar_cpf(self, e):
            cpf = ''.join(filter(str.isdigit, self.cpfclb.value))[:11]
            self.cpfclb.value = cpf
            self.cpfclb.update()

        def formatar_cpf(self, e):
            cpf = ''.join(filter(str.isdigit, self.cpfclb.value))[:11]
            if len(cpf) == 11:
                self.cpfclb.value = f"{cpf[:3]}.{cpf[3:6]}.{cpf[6:9]}-{cpf[9:]}"
                self.cpfclb.update()

        def limitar_data(self, e):
            data = ''.join(filter(str.isdigit, self.datanascimentoclb.value))[:8]
            self.datanascimentoclb.value = data
            self.datanascimentoclb.update()

        def formatar_data(self, e):
            data = ''.join(filter(str.isdigit, self.datanascimentoclb.value))[:8]
            if len(data) == 8:
                self.datanascimentoclb.value = f"{data[:2]}/{data[2:4]}/{data[4:]}"
                self.datanascimentoclb.update()
                
        def on_resize(self,e):            
            if self.page.route == "/gerardoc":
                self.responsive = Responsive(self.page)
                self.responsive.atualizar_widgets(self.build_view())

        def build_view(self):
            if self.responsive.is_desktop():
                self.listview = ft.ListView(expand=True)
                self.listviewtypes = ft.ListView(expand=True)
                self.listviewexam = ft.ListView(expand=True)
                self.checkbox = ft.Checkbox(label=ft.Text("Gerar em todos os modelos"),on_change=self.selecionar_todos,
                                            label_position=ft.LabelPosition.LEFT,check_color="#26BD00",active_color="#D3FACA")
                self.drop = ft.Dropdown(label="Empresas",width=200)

                

            
                self.date = ft.TextField(label="Data do exame",border_radius=16,width=140)
                self.nomeclb = ft.TextField(label="Nome Completo",border_radius=16,width=250)
                self.cpfclb = ft.TextField(label="CPF",border_radius=16,width=250, on_change=self.limitar_cpf, on_blur=self.formatar_cpf)
                self.datanascimentoclb = ft.TextField(label="Data de nascimento",border_radius=16,width=250, on_change=self.limitar_data, on_blur=self.formatar_data)
                self.funcaoclb = ft.TextField(label="Função",border_radius=16,width=250)
                self.setorclb = ft.TextField(label="Setor",border_radius=16,width=250)
                self.clbinterface =ft.Column([
                        ft.Row([
                            ft.Text("Insira os dados do colaborador")
                        ],alignment=ft.MainAxisAlignment.CENTER,vertical_alignment=ft.CrossAxisAlignment.CENTER
                        ),
                        ft.Row([
                            self.nomeclb,self.cpfclb
                        ],alignment=ft.MainAxisAlignment.CENTER,vertical_alignment=ft.CrossAxisAlignment.CENTER
                        ),
                        ft.Row([
                            self.datanascimentoclb
                        ],alignment=ft.MainAxisAlignment.CENTER,vertical_alignment=ft.CrossAxisAlignment.CENTER
                        ),
                        ft.Row([
                            self.setorclb,self.funcaoclb
                        ],alignment=ft.MainAxisAlignment.CENTER,vertical_alignment=ft.CrossAxisAlignment.CENTER
                        ),
                        ft.Row([
                            ft.TextButton(text="Gerar documento",icon=ft.Icons.SEND_AND_ARCHIVE,
                                            style=ft.ButtonStyle(color=ft.Colors.BLACK,elevation=5,bgcolor=ft.Colors.LIGHT_GREEN_ACCENT,text_style=ft.TextStyle(weight=ft.FontWeight.W_400,color=ft.Colors.WHITE)),
                                            icon_color=ft.Colors.BLACK45,width=150,on_click=self.gerar_documento),
                            ft.TextButton(text="Limpar",icon=ft.Icons.SEND_AND_ARCHIVE,
                                            style=ft.ButtonStyle(color=ft.Colors.BLACK,elevation=5,bgcolor=ft.Colors.LIGHT_GREEN_ACCENT,text_style=ft.TextStyle(weight=ft.FontWeight.W_400,color=ft.Colors.WHITE)),
                                            icon_color=ft.Colors.BLACK45,on_click=lambda e: self.limpar(e),width=150)
                        ],alignment=ft.MainAxisAlignment.CENTER,vertical_alignment=ft.CrossAxisAlignment.CENTER
                        ),
                    ],alignment=ft.MainAxisAlignment.CENTER,horizontal_alignment=ft.CrossAxisAlignment.CENTER)
                self.checkdate = ft.Checkbox(label="Usar data de hoje",check_color=ft.Colors.BLACK,active_color="#74FE4E",on_change=lambda e: self.selectdate(e))
                self.risk_selector = ft.ElevatedButton(text="Defina os Riscos",icon=ft.Icons.WARNING,icon_color=ft.Colors.YELLOW_800,color=ft.Colors.BLACK54,on_click=lambda e: self.risk_page(e))
                self.dates = ft.Container(
                    content=ft.Column([
                        ft.Column([self.checkdate,self.date])
                    ])
                )
                self.riscos = ft.Container(
                    content=ft.Column([
                        ft.Column([self.risk_selector])
                    ])
                )
                self.view_exames = ft.Container(content=ft.Column([
                    self.listviewexam
                ],scroll=ft.ScrollMode.ALWAYS,expand=True),adaptive=True,expand=True)
                self.doccontent = ft.Column([
                    ft.Row([ft.Text("Gerar documento", size=30)],alignment=ft.MainAxisAlignment.CENTER,vertical_alignment=ft.CrossAxisAlignment.CENTER),
                    ft.Container(
                        content=ft.Row([
                            ft.Column([
                                self.main.cardmain("Modelos de Exame",self.page.width*0.19,None,self.listview,False),
                                self.main.cardmain("Tipos do exame",self.page.width*0.19,None,self.listviewtypes,True)
                            ],alignment=ft.MainAxisAlignment.CENTER,horizontal_alignment=ft.CrossAxisAlignment.CENTER
                            ),
                            ft.Column([
                                ft.Row([self.main.cardmain("Riscos\nOcupacionais".upper(),None,None,self.riscos,True),
                                        self.main.cardmain("Contratante".upper(),None,None,self.drop,True),
                                        self.main.cardmain("Data a elencar".upper(),None,None,self.dates,True)],expand=True),
                                self.main.cardmain("Colaborador".upper(),None,None,self.clbinterface,False)
                            ],alignment=ft.MainAxisAlignment.CENTER,horizontal_alignment=ft.CrossAxisAlignment.CENTER
                            ),
                            ft.Column([
                                self.main.cardmain("Exames a elencar",self.page.width*0.20,self.page.height*0.70,self.view_exames,False)
                            ],alignment=ft.MainAxisAlignment.CENTER,horizontal_alignment=ft.CrossAxisAlignment.CENTER
                            ),
                        ],alignment=ft.MainAxisAlignment.CENTER,vertical_alignment=ft.CrossAxisAlignment.CENTER,spacing=5),
                        margin=ft.margin.only(top=-20)
                )],scroll=ft.ScrollMode.ADAPTIVE)
                self.listview.controls.append(ft.Text("Modelos disponiveis: ",text_align=ft.TextAlign.START))
                self.listviewtypes.controls.append(ft.Text("Qual tipo do documento: ",text_align=ft.TextAlign.START))
                self.listviewexam.controls.append(ft.Text("Quais exames elencar: ",text_align=ft.TextAlign.START))
                self.listview.controls.append(self.checkbox)
                self.page.run_task(self.atualizar_lista_modelos)
                self.page.run_task(self.carregardrop)
                return ft.Row(
                    [
                        ft.Column([self.sidebar.build()],alignment=ft.MainAxisAlignment.START,horizontal_alignment=ft.CrossAxisAlignment.START),
                        ft.Container(content=self.doccontent,expand=True,adaptive=True)
                    ],
                    width=self.page.width,
                    height=self.page.height,

                
                )
            elif self.responsive.is_tablet():
                self.listview = ft.ListView(expand=True)
                self.listviewtypes = ft.ListView(expand=True)
                self.listviewexam = ft.ListView(expand=True)
                self.checkbox = ft.Checkbox(label=ft.Text("Gerar em todos os modelos"),on_change=self.selecionar_todos,
                                            label_position=ft.LabelPosition.LEFT,check_color="#26BD00",active_color="#D3FACA")
                self.drop = ft.Dropdown(label="Empresas",autofocus=True,options=[],width=200)
                self.date = ft.TextField(label="Data do exame",border_radius=16,width=140)
                self.nomeclb = ft.TextField(label="Nome Completo",border_radius=16,width=300)
                self.cpfclb = ft.TextField(label="CPF",border_radius=16,width=300)
                self.datanascimentoclb = ft.TextField(label="Data de nascimento",border_radius=16,width=300)
                self.funcaoclb = ft.TextField(label="Função",border_radius=16,width=300)
                self.setorclb = ft.TextField(label="Setor",border_radius=16,width=300)
                self.clbinterface =ft.Column([
                        ft.Row([
                            ft.Text("Insira os dados do colaborador")
                        ],alignment=ft.MainAxisAlignment.CENTER,vertical_alignment=ft.CrossAxisAlignment.CENTER
                        ),
                        ft.Row([
                            self.nomeclb,self.cpfclb
                        ],alignment=ft.MainAxisAlignment.CENTER,vertical_alignment=ft.CrossAxisAlignment.CENTER
                        ),
                        ft.Row([
                            self.datanascimentoclb
                        ],alignment=ft.MainAxisAlignment.CENTER,vertical_alignment=ft.CrossAxisAlignment.CENTER
                        ),
                        ft.Row([
                            self.setorclb,self.funcaoclb
                        ],alignment=ft.MainAxisAlignment.CENTER,vertical_alignment=ft.CrossAxisAlignment.CENTER
                        ),
                        ft.Row([
                            ft.TextButton(text="Gerar documento",icon=ft.Icons.SEND_AND_ARCHIVE,
                                            style=ft.ButtonStyle(color=ft.Colors.BLACK,elevation=5,bgcolor=ft.Colors.LIGHT_GREEN_ACCENT,text_style=ft.TextStyle(weight=ft.FontWeight.W_400,color=ft.Colors.WHITE)),
                                            icon_color=ft.Colors.BLACK45,width=150,on_click=self.gerar_documento),
                            ft.TextButton(text="Limpar",icon=ft.Icons.SEND_AND_ARCHIVE,
                                            style=ft.ButtonStyle(color=ft.Colors.BLACK,elevation=5,bgcolor=ft.Colors.LIGHT_GREEN_ACCENT,text_style=ft.TextStyle(weight=ft.FontWeight.W_400,color=ft.Colors.WHITE)),
                                            icon_color=ft.Colors.BLACK45,on_click=lambda e: self.limpar(e),width=150)
                        ],alignment=ft.MainAxisAlignment.CENTER,vertical_alignment=ft.CrossAxisAlignment.CENTER
                        ),
                    ],alignment=ft.MainAxisAlignment.CENTER,horizontal_alignment=ft.CrossAxisAlignment.CENTER)
                self.checkdate = ft.Checkbox(label="Usar data de hoje",check_color=ft.Colors.BLACK,active_color="#74FE4E",on_change=lambda e: self.selectdate(e))
                self.risk_selector = ft.ElevatedButton(text="Defina os Riscos",icon=ft.Icons.WARNING,icon_color=ft.Colors.YELLOW_800,color=ft.Colors.BLACK54,on_click=lambda e: self.risk_page(e))
                self.dates = ft.Container(
                    content=ft.Column([
                        ft.Column([self.checkdate,self.date])
                    ])
                )
                self.riscos = ft.Container(
                    content=ft.Column([
                        ft.Column([self.risk_selector])
                    ])
                )
                self.doccontent = ft.Column([
                    ft.Row([ft.Text("Gerar documento", size=30)],alignment=ft.MainAxisAlignment.CENTER,vertical_alignment=ft.CrossAxisAlignment.CENTER),
                    ft.Container(
                        content=ft.Row([
                            ft.Column([
                                self.main.cardmain("Modelos de Exame",self.page.width*0.19,None,self.listview,False),
                                self.main.cardmain("Tipos do exame",self.page.width*0.19,None,self.listviewtypes,True)
                            ],alignment=ft.MainAxisAlignment.START,horizontal_alignment=ft.CrossAxisAlignment.CENTER
                            ),
                            ft.Column([
                                ft.Row([self.main.cardmain("Riscos\nOcupacionais".upper(),None,None,self.riscos,True),
                                        self.main.cardmain("Contratante".upper(),None,None,self.drop,True),
                                        self.main.cardmain("Data a elencar".upper(),None,None,self.dates,True)],expand=True),
                                self.main.cardmain("Colaborador".upper(),None,None,self.clbinterface,False)
                            ],alignment=ft.MainAxisAlignment.CENTER,horizontal_alignment=ft.CrossAxisAlignment.CENTER
                            ),
                            ft.Column([
                                self.main.cardmain("Exames a elencar",self.page.width*0.20,None,self.listviewexam,False)
                            ],alignment=ft.MainAxisAlignment.CENTER,horizontal_alignment=ft.CrossAxisAlignment.CENTER
                            ),
                        ],alignment=ft.MainAxisAlignment.CENTER,vertical_alignment=ft.CrossAxisAlignment.CENTER,spacing=35)
                )])
                self.listview.controls.append(ft.Text("Modelos disponiveis: ",text_align=ft.TextAlign.START))
                self.listviewtypes.controls.append(ft.Text("Qual tipo do documento: ",text_align=ft.TextAlign.START))
                self.listviewexam.controls.append(ft.Text("Quais exames elencar: ",text_align=ft.TextAlign.START))
                self.listview.controls.append(self.checkbox)
                self.page.run_task(self.atualizar_lista_modelos)
                self.page.run_task(self.carregardrop)
                return ft.Column(
                    [
                        ft.Row([self.sidebar.build()],alignment=ft.MainAxisAlignment.START,vertical_alignment=ft.CrossAxisAlignment.START),
                        ft.Container(content=self.doccontent,expand=True,adaptive=True)
                    ],
                    width=self.page.width,
                    height=self.page.height,
                    scroll=ft.ScrollMode.ADAPTIVE
                )
            elif self.responsive.is_mobile():
                return ft.Column(
                    [
                        ft.Row([self.sidebar.build()],alignment=ft.MainAxisAlignment.START,vertical_alignment=ft.CrossAxisAlignment.START),
                        ft.Container(content=ft.Text("Uso não permitido",size=35),expand=True)
                    ],
                    width=self.page.width,
                    height=self.page.height
                )
        
        def limpar(self,e):
            self.nomeclb.value = ""
            self.datanascimentoclb.value = ""
            self.cpfclb.value = ""
            self.setorclb.value = ""
            self.funcaoclb.value = ""
            self.nomeclb.update()
            self.datanascimentoclb.update()
            self.cpfclb.update()
            self.setorclb.update()
            self.funcaoclb.update()
        
        def carregar_modelos_excel(self):
            modelos = []
            modelos_dir = Path(r"modelos_excel")
                
            if not modelos_dir.exists():
                 modelos_dir.mkdir()
                
            for arquivo in modelos_dir.glob("*.xlsx"):
                modelos.append(arquivo.name.replace(".xlsx",""))
            return modelos           
        
        async def atualizar_lista_modelos(self):
            self.listview.controls.clear()
            self.listviewexam.controls.clear()
            self.listviewtypes.controls.clear()
            self.types = [
                "ADMISSIONAL",
                "DEMISSIONAL" ,
                "MUDANÇA DE RISCO",
                "PERIÓDICO",
                "RETORNO A FUNÇÃO"
            ]
            self.exams =[
                "ANAMNESE",
                "AUDIOMETRIA",
                "HEMOGRAMA",
                "PARASITOLOGICO",
                "BRUCELOSE",
                "ELETROCARDIOGRAMA",
                "ELETROENCEFALOGRAMA",
                "GLICEMIA",
                "PSICOSSOCIAL",
                "V.D.R.L",
                "ESPIROMETRIA",
                "RAIO-X (TORAX)",
                "RAIO-X (LOMBO)",
                "ACUIDADE VISUAL",
                "TESTE ROMBERG",
                "TGO (Aspartato)",
                "TGP (Alanina)",
                "CREATININA",
                "URÉIA",
                "TOXICOLOGICO"
            ]
            for modelo in self.modelos_excel:
                self.listview.controls.append(
                    ft.ListTile(
                        title=ft.Text(modelo),
                        on_click=self.selecionar_modelo,
                        enable_feedback=True,
                        selected_color="#26BD00",
                        title_alignment=ft.ListTileTitleAlignment.CENTER,
                        bgcolor=ft.Colors.GREY_100,
                        selected_tile_color=ft.Colors.with_opacity(0.2,ft.Colors.LIGHT_GREEN_ACCENT_100)
                            
                    )
                )
            if not self.modelos_excel:
                self.listview.controls.append(
                    ft.ListTile(
                        title=ft.Text("Nenhum modelo encontrado na pasta 'modelos_excel'"),
                        subtitle=ft.Text("Adicione modelos Excel (.xlsx) nesta pasta")
                    )
                )
            for tipo in self.types:
                self.listviewtypes.controls.append(
                    ft.ListTile(
                        title=ft.Text(tipo),
                        enable_feedback=True,
                        on_click=self.selecionar_tipo,
                        selected_color="#26BD00",
                        title_alignment=ft.ListTileTitleAlignment.CENTER,
                        bgcolor=ft.Colors.with_opacity(0.6,ft.Colors.GREY_200),
                        height=40,
                        selected_tile_color=ft.Colors.with_opacity(0.2,ft.Colors.LIGHT_GREEN_ACCENT_100)
                    )
                )
            for exam in self.exams:
                self.listviewexam.controls.append(
                    ft.ListTile(
                        title=ft.Text(exam),
                        enable_feedback=True,
                        on_click=self.selecionar_exames,
                        selected_color="#26BD00",
                        title_alignment=ft.ListTileTitleAlignment.CENTER,
                        bgcolor=ft.Colors.with_opacity(0.6,ft.Colors.GREY_200),
                        height=40,
                        selected_tile_color=ft.Colors.with_opacity(0.2,ft.Colors.LIGHT_GREEN_ACCENT_100)
                    )
                )

        async def selecionar_modelo(self, e):
            e.control.selected = not getattr(e.control, 'selected', False)
            tiles = [control for control in self.listview.controls if isinstance(control, ft.ListTile)]
            todos_selecionados = all(getattr(tile, 'selected', False) for tile in tiles)
            self.checkbox.value = todos_selecionados
            self.page.update()
        
        async def selecionar_tipo(self, e):
            if getattr(e.control, 'selected', False):
                for control in self.listviewtypes.controls:
                    control.selected = False # type: ignore
            else:
                for control in self.listviewtypes.controls:
                    control.selected = False # type: ignore
                e.control.selected = True
            self.page.update()           
        
        async def selecionar_todos(self,e):
            selecionar_todos = e.control.value
                # Aplica a seleção a todos os ListTiles
            for control in self.listview.controls:
                if isinstance(control, ft.ListTile):
                    control.selected = selecionar_todos       
        
        async def selecionar_exames(self, e):
            e.control.selected = not getattr(e.control, 'selected', False)
            self.page.update()
        
        async def carregardrop(self):
            self.drop.options.clear() #type: ignore
            if self.empresas_drop:
                for empresa in self.empresas_drop:
                    self.drop.options.append( #type: ignore
                        ft.DropdownOption(
                            text=f"• {empresa[0]}", style=ft.TextStyle(size=10)
                        )
                    ) 
            else:
                self.drop.options.append( #type: ignore
                    ft.DropdownOption(
                        text=f"{"Nenhuma Empresa Cadastrada"}",
                        disabled=True
                    )
                )            
    
        def risk_page(self,e):
            def close(e):
                self.page.close(modal)
            def catch_risk(e):
                if not e.control.label in self.risk:
                    self.risk.append(e.control.label)
                    match e.control.data:
                        case "f":
                            self.risk_fisico.append(e.control.label)
                        case "q":
                            self.risk_quimico.append(e.control.label)
                        case "b":
                            self.risk_biologico.append(e.control.label)
                        case "e":
                            self.risk_ergonomico.append(e.control.label)
                if e.control.value is False:
                    self.risk.remove(e.control.label)
                    match e.control.data:
                        case "f":
                            self.risk_fisico.remove(e.control.label)
                        case "q":
                            self.risk_quimico.remove(e.control.label)
                        case "b":
                            self.risk_biologico.remove(e.control.label)
                        case "e":
                            self.risk_ergonomico.remove(e.control.label)
            self.drop.update()
            def criar_risco(risco,tipo):
                rsk = ft.Checkbox(label=risco,data=tipo,on_change=catch_risk)
                return rsk
            riscos_fisicos = [
                "Calor",
                "Frio",
                "Radiações-ION",
                "Radiações-Não-ION",
                "Ruidos",
                "Vibrações",
            ]
            riscos_quimicos = [
                "Fumos Metálicos",
                "Gases",
                "Hidrocarbonetos",
                "Neblinas",
                "Produtos Tóxicos",
                "Névoas",
                "Solventes",
                "Vapores Orgânicos",
                "Poeiras Minerais",]
            riscos_biologicos = [
                "Bactérias",
                "Fungos",
                "Parásitas",
                "Vírus",
                "Bacilos",
                "Protozoarios",
                "Micoses",]
            riscos_ergonomicos = [
                "Esforço Fisícos",
                "Levantamento de carga",
                "Impacto de Objetos",
                "Movimento Repetitivo",
                "Trabalho em Turno",
                "Postura Inadequada",
                "Objetos Perfurocortantes",
                "Deslocamento em Ambiente Industrial",]
            modal = ft.AlertDialog(
                modal=True,
                bgcolor="#FFFFFF",
                title="Riscos Ocupacionais",
                actions=[ft.ElevatedButton("Enviar",on_click=close)],
                content=ft.Column(
                    [
                        ft.Text("Selecione os Riscos Ocupacionais: "),
                        ft.Row([
                            ft.Column([
                                ft.Text("FISÍCOS",weight=ft.FontWeight.BOLD,size=20),
                                criar_risco(riscos_fisicos[0],"f"),
                                criar_risco(riscos_fisicos[1],"f"),
                                criar_risco(riscos_fisicos[2],"f"),
                                criar_risco(riscos_fisicos[3],"f"),
                                criar_risco(riscos_fisicos[4],"f"),
                                criar_risco(riscos_fisicos[5],"f"),
                                ],alignment=ft.MainAxisAlignment.CENTER,horizontal_alignment=ft.CrossAxisAlignment.START),
                            ft.Column([
                                ft.Text("QUÍMICOS",weight=ft.FontWeight.BOLD,size=20),
                                criar_risco(riscos_quimicos[0],"q"),
                                criar_risco(riscos_quimicos[1],"q"),
                                criar_risco(riscos_quimicos[2],"q"),
                                criar_risco(riscos_quimicos[3],"q"),
                                criar_risco(riscos_quimicos[4],"q"),
                                criar_risco(riscos_quimicos[5],"q"),
                                criar_risco(riscos_quimicos[6],"q"),
                                criar_risco(riscos_quimicos[7],"q"),
                                criar_risco(riscos_quimicos[8],"q"),

                            ],horizontal_alignment=ft.CrossAxisAlignment.START),
                            ft.Column([
                                ft.Text("BIOLÓGICOS",weight=ft.FontWeight.BOLD,size=20),
                                criar_risco(riscos_biologicos[0],"b"),
                                criar_risco(riscos_biologicos[1],"b"),
                                criar_risco(riscos_biologicos[2],"b"),
                                criar_risco(riscos_biologicos[3],"b"),
                                criar_risco(riscos_biologicos[4],"b"),
                                criar_risco(riscos_biologicos[5],"b"),
                                criar_risco(riscos_biologicos[6],"b"),
                            ],horizontal_alignment=ft.CrossAxisAlignment.START),
                            ft.Column([
                                ft.Text("ERGONÔMICOS",weight=ft.FontWeight.BOLD,size=20),
                                criar_risco(riscos_ergonomicos[0],"e"),
                                criar_risco(riscos_ergonomicos[1],"e"),
                                criar_risco(riscos_ergonomicos[2],"e"),
                                criar_risco(riscos_ergonomicos[3],"e"),
                                criar_risco(riscos_ergonomicos[4],"e"),
                                criar_risco(riscos_ergonomicos[5],"e"),
                                criar_risco(riscos_ergonomicos[6],"e"),
                                criar_risco(riscos_ergonomicos[7],"e"),
                            ],horizontal_alignment=ft.CrossAxisAlignment.START),
                        ],alignment=ft.MainAxisAlignment.CENTER,vertical_alignment=ft.CrossAxisAlignment.START,spacing=100)
                        
                    ],height=500,width=1100
                ),
            )
            self.page.open(modal)

        def clean_risks(self):
            self.risk = []
            self.risk_fisico = []
            self.risk_quimico = []
            self.risk_biologico = []
            self.risk_ergonomico = []

        def calcular_idade(self,nascimento: date) -> int:
            hoje = date.today()
            idade = hoje.year - nascimento.year
            if (hoje.month, hoje.day) < (nascimento.month, nascimento.day):
                idade -= 1
            return idade

        def selectdate(self, e):
            if self.date.disabled == False:
                self.date.disabled = True
                self.date.update()
                self.dataselect = dt.datetime.now().strftime("%d/%m/%Y")
            else:
                self.date.disabled = False
                self.dataselect = None
                self.date.update()

        def gerar_documento(self, e):            
            if not self.checkdate.value:
                self.dataselect = self.date.value
            self.empresas = puxardados(self.drop.value)
            nome = self.nomeclb.value if self.nomeclb.value else None
            cpf = self.cpfclb.value if self.cpfclb.value else None
            nascimento = self.datanascimentoclb.value if self.datanascimentoclb.value else None
            funcao = self.funcaoclb.value if self.funcaoclb.value else None
            setor = self.setorclb.value if self.setorclb.value else None
            modelos_selecionados = [
                control.title.value # type: ignore
                for control in self.listview.controls 
                if isinstance(control, ft.ListTile) and getattr(control, "selected", False)]
            exames_selecionados = [
                control.title.value # type: ignore
                for control in self.listviewexam.controls 
                if isinstance(control, ft.ListTile) and getattr(control, "selected", False)]
            tipo_exame = [
                control.title.value # type: ignore
                for control in self.listviewtypes.controls 
                if isinstance(control, ft.ListTile) and getattr(control, "selected", False)]
            dtn = nascimento.replace("/","-")
            idade = dt.datetime.strptime(dtn, "%d-%m-%Y").date()
            idade = self.calcular_idade(idade)
            if not self.empresas:
                self.main.barra_aviso("Selecione pelo menos uma empresa!",ft.Colors.YELLOW,text_color=ft.Colors.BLACK)
                return
            if not modelos_selecionados:
                self.main.barra_aviso("Selecione pelo menos um modelo!",ft.Colors.YELLOW,text_color=ft.Colors.BLACK)
                return
            if not tipo_exame:
                self.main.barra_aviso("Selecione o tipo do exame!",ft.Colors.YELLOW,text_color=ft.Colors.BLACK)
                return
            if nome == None or cpf == None or nascimento == None:
                self.main.barra_aviso("Preencha todos os campos obrigatórios!",ft.Colors.YELLOW,text_color=ft.Colors.BLACK)
                return
            if not self.dataselect:
                self.main.barra_aviso("Selecione a data do exame!",ft.Colors.YELLOW,text_color=ft.Colors.BLACK)
            if nome and cpf and nascimento:
                if setor == None or funcao == None:
                    self.main.barra_aviso("AVISO: Exame sendo gerado sem função ou setor!",ft.Colors.YELLOW,text_color=ft.Colors.BLACK)
                for modelo in modelos_selecionados:
                    caminho_modelo = Path(r"modelos_excel") / f"{modelo}.xlsx"
                    if not caminho_modelo.exists():
                        continue  
                    if modelo == "ANAMNESE":
                        wb = load_workbook(caminho_modelo)
                        ws = wb.active
                        ws["C10"] = nome # type: ignore
                        ws["C11"] = cpf # type: ignore
                        ws["E12"] = nascimento # type: ignore
                        ws["D13"] = funcao # type: ignore
                        ws["I13"] = setor # type: ignore
                        ws["I12"] = f"{idade} Anos"
                        ws["G11"] = self.empresas[0] # type: ignore
                        ws["G48"] = self.dataselect # type: ignore
                        ws["E9"] = tipo_exame[0]
                        saida = Path(r"documentos_gerados")
                        saida.mkdir(exist_ok=True)
                        self.nome_arquivo = f"{modelo} {self.empresas[0].replace(' ', '-')} {nome.replace(' ', '-')} {dt.datetime.now().strftime('%d-%m-%Y %H-%M')} .xlsx"
                        wb.save(saida / self.nome_arquivo)
                    elif modelo == "ASO":
                        if not self.risk:
                            self.main.barra_aviso("Selecione os riscos ocupacionais!", ft.Colors.YELLOW, text_color=ft.Colors.BLACK)
                            return
                        if not exames_selecionados:
                            self.main.modal("Aviso","Exame será gerado sem exames complementares!")
                        wb = load_workbook(caminho_modelo)
                        ws = wb.active
                        ws["D13"] = nome# type: ignore
                        ws["J14"] = cpf # type: ignore
                        ws["D14"] = nascimento # type: ignore
                        ws["D15"] = funcao # type: ignore
                        ws["J15"] = setor # type: ignore
                        ws["J13"] = f"{idade} Anos"
                        ws["D9"] = self.empresas[0] # type: ignore
                        ws["L6"] = self.dataselect # type: ignore
                        ws["J9"] = self.empresas[1] # type: ignore
                        ws["D10"] = self.empresas[2]
                        ws["J10"] = self.empresas[3]
                        for i, valor in enumerate(exames_selecionados):
                            ws[f"C{31+i}"] = valor
                            ws[f"E{31+i}"] = self.dataselect
                        ws["G6"] = tipo_exame[0]
                        for i, valor in enumerate(self.risk_fisico):
                            ws[f"D{19+i}"] = valor
                        for i, valor in enumerate(self.risk_quimico):
                            ws[f"E{19+i}"] = valor
                        for i, valor in enumerate(self.risk_biologico):
                            ws[f"G{19+i}"] = valor
                        for i, valor in enumerate(self.risk_ergonomico):
                            ws[f"I{19+i}"] = valor
                        saida = Path(r"documentos_gerados")
                        saida.mkdir(exist_ok=True)
                        self.nome_arquivo = f"{modelo} {self.empresas[0].replace(' ', '-')} {nome.replace(' ', '-')} {dt.datetime.now().strftime('%d-%m-%Y %H-%M')}.xlsx"
                        wb.save(saida / self.nome_arquivo)
                    elif modelo == "AUDIOMETRIA":
                        wb = load_workbook(caminho_modelo)
                        ws = wb.active
                        ws["E15"] = nome# type: ignore
                        ws["L15"] = cpf # type: ignore
                        ws["P16"] = nascimento # type: ignore
                        ws["E16"] = funcao # type: ignore
                        ws["E14"] = self.empresas[0] # type: ignore
                        ws["N17"] = self.dataselect # type: ignore
                        ws["M14"] = self.empresas[1] # type: ignore
                        ws["L16"] = f"{idade} Anos"
                        ws["H11"] = tipo_exame[0]
                        saida = Path(r"documentos_gerados")
                        saida.mkdir(exist_ok=True)
                        self.nome_arquivo = f"{modelo} {self.empresas[0].replace(' ', '-')} {nome.replace(' ', '-')} {dt.datetime.now().strftime('%d-%m-%Y %H-%M')} .xlsx"
                        wb.save(saida / self.nome_arquivo)
                self.nomeclb.value = None
                self.cpfclb.value = None
                self.datanascimentoclb.value = None
                self.funcaoclb.value = None
                self.setorclb.value = None
                self.nomeclb.update()
                self.cpfclb.update()
                self.datanascimentoclb.update()
                self.funcaoclb.update()
                self.setorclb.update()
                self.modal = ft.SnackBar(content=ft.Text("Documentos Gerados"),bgcolor=ft.Colors.GREEN)
                self.page.snack_bar = self.modal
                self.modal.open = True
                self.clean_risks()
                self.page.add(self.modal)      
                
        def show_loading(self,page, show=True):
            loading = ft.Container(
                content=ft.Column(
                    [
                        ft.ProgressRing(width=50, height=50, color=ft.Colors.LIGHT_GREEN),
                        ft.Text("Carregando...", size=16, color=ft.Colors.BLACK)
                    ],
                    alignment=ft.MainAxisAlignment.CENTER,
                    horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                    spacing=20
                ),
                alignment=ft.Alignment(x=0,y=0),
                bgcolor=ft.Colors.with_opacity(0.4, ft.Colors.WHITE30),
                expand=True
            )
                
            if show:
                page.overlay.append(loading)
            else:
                if loading in page.overlay:
                    page.overlay.remove(loading)
                    page.overlay.clear()
                    loading.visible = False
                    page.update()
                else:
                    pass    
