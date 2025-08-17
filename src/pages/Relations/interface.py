import flet as ft
from pathlib import Path
from openpyxl import load_workbook
from src.utils.telaresize import Responsive
import re
from datetime import datetime

class Relations:
    def __init__(self, page: ft.Page):
        self.page = page
        self.responsive = Responsive(self.page)
        self.empresas_exames = self.listar_empresas()
        self.search_field = ft.TextField(
            hint_text="Pesquisar empresa",
            on_change=self.filtrar_empresas,
            width=400,
            prefix_icon=ft.Icons.SEARCH,
            border_color=ft.Colors.GREY_400,
            bgcolor=ft.Colors.WHITE,
            color=ft.Colors.GREY_900,
            border_radius=8,
            height=48
        )
        self.expansion_list = ft.ExpansionPanelList(
            expand_icon_color=ft.Colors.GREY_700,
            on_change=self.expandir_empresa
        )
        self.carregar_empresas()

    def tratar_data(self,data_exame):
        if not data_exame:
            return None
        if isinstance(data_exame, datetime):
            return data_exame.strftime("%d/%m/%Y")
        if isinstance(data_exame, int) or isinstance(data_exame, float):
            data_str = str(int(data_exame))
            try:
                if len(data_str) == 8:
                    return datetime.strptime(data_str, "%d%m%Y").strftime("%d/%m/%Y")
            except:
                return data_str
        if isinstance(data_exame, str):
            for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d%m%Y"):
                try:
                    return datetime.strptime(data_exame, fmt).strftime("%d/%m/%Y")
                except:
                    continue
            return data_exame
        return str(data_exame)

    def limpar_nome(self, nome):
        """Remove caracteres inválidos para nomes de arquivo no Windows"""
        return re.sub(r'[\\/*?:"<>|]', "", nome)

    # Agora só lista empresas, não carrega exames
    def listar_empresas(self):
        relacoes_dir = Path("relacoes")
        empresas = {}

        if not relacoes_dir.exists():
            return {}

        for arquivo in relacoes_dir.rglob("*.[xX][lL][sS][xX]"):
            if "modelo" in arquivo.name.lower():
                continue

            partes = arquivo.stem.split(" - ")
            if len(partes) >= 3:
                empresa_nome = partes[0].strip()
                cnpj = ""
            else:
                empresa_nome = arquivo.stem.split("_")[0]
                cnpj = ""

            empresas[(empresa_nome, cnpj)] = arquivo

        return empresas

    # Carrega exames só quando abrir empresa
    def carregar_exames_empresa(self, arquivo):
        exames = []
        try:
            wb = load_workbook(arquivo)
            ws = wb.active
            for row in ws.iter_rows(min_row=6, values_only=True):
                if not row or len(row) < 11:
                    continue
                nome_colaborador = row[1]
                exames_nome = row[3]
                data_exame = self.tratar_data(row[10])
                if nome_colaborador or exames_nome or data_exame:
                    exames.append({
                        "exame": exames_nome,
                        "colaborador": nome_colaborador,
                        "data": data_exame,
                        "hora": "",
                        "arquivo": arquivo
                    })
        except Exception as e:
            print(f"Erro ao ler {arquivo}: {e}")
        return exames

    def carregar_empresas(self, filtro=""):
        self.expansion_list.controls.clear()
        for (empresa, cnpj), arquivo in self.empresas_exames.items():
            if filtro.lower() in empresa.lower():
                if len(self.expansion_list.controls) >= 50:  # limite de empresas na tela
                    break
                self.expansion_list.controls.append(
                    self.criar_painel_empresa(empresa, cnpj, arquivo)
                )
        self.page.update()

    def filtrar_empresas(self, e):
        filtro = e.control.value
        self.carregar_empresas(filtro)

    def excluir_exame(self, arquivo, colaborador, exame_texto, data):
        try:
            wb = load_workbook(arquivo)
            ws = wb.active
            for row_num in range(6, ws.max_row + 1):
                if (ws[f"B{row_num}"].value == colaborador and
                    ws[f"D{row_num}"].value == exame_texto and
                    str(ws[f"K{row_num}"].value) == str(data)):
                    ws[f"B{row_num}"] = None
                    ws[f"D{row_num}"] = None
                    ws[f"K{row_num}"] = None
                    break
            wb.save(arquivo)
            print(f"Exame excluído com sucesso")
        except Exception as e:
            print(f"Erro ao excluir exame: {e}")

        # Recarrega só a empresa modificada
        self.empresas_exames = self.listar_empresas()
        self.carregar_empresas(self.search_field.value if hasattr(self, 'search_field') else "")

    def criar_painel_empresa(self, empresa, cnpj, arquivo):
        return ft.ExpansionPanel(
            can_tap_header=True,
            header=ft.Container(
                content=ft.Column(
                    [
                        ft.Text(
                            empresa,
                            size=18,
                            weight=ft.FontWeight.NORMAL,
                            color=ft.Colors.GREY_900,
                            font_family="Roboto"
                        ),
                        ft.Text(
                            cnpj,
                            size=14,
                            color=ft.Colors.GREY_600,
                            font_family="Roboto"
                        ) if cnpj else ft.Container()
                    ],
                    spacing=2
                ),
                padding=ft.padding.symmetric(vertical=8, horizontal=12),
                bgcolor=ft.Colors.with_opacity(0.08, ft.Colors.GREY_100), 
                border_radius=20,
            ),
            content=ft.Container(
                content=ft.Text(
                    "Clique para carregar exames...",
                    italic=True,
                    size=14,
                    color=ft.Colors.GREY_500
                ),
                bgcolor=ft.Colors.WHITE,
                border_radius=20,
                padding=16
            )
        )

    def expandir_empresa(self, e):
        index = int(e.data)  # índice do painel expandido
        if index < 0:
            return  # significa que fechou tudo

        panel = self.expansion_list.controls[index]
        empresa, cnpj = list(self.empresas_exames.keys())[index]
        arquivo = self.empresas_exames[(empresa, cnpj)]

        exames = self.carregar_exames_empresa(arquivo)

        panel.content = ft.Container(
            content=ft.Column(
                [
                    self.criar_linha_exame(exame) for exame in exames
                ] if exames else [
                    ft.Text(
                        "Nenhum exame realizado.",
                        italic=True,
                        size=14,
                        color=ft.Colors.GREY_500,
                        font_family="Roboto"
                    )
                ],
                spacing=4
            ),
            bgcolor=ft.Colors.WHITE,
            border_radius=20,
            padding=16
        )
        self.page.update()

    def criar_linha_exame(self, exame):
        return ft.Row(
            [
                ft.Text(
                    f"Colaborador: {exame['colaborador']} | Exame: {exame['exame']} | Data: {exame['data']} {exame['hora']}",
                    size=15,
                    color=ft.Colors.GREY_900,
                    font_family="Roboto",
                    expand=True
                ),
                ft.IconButton(
                    icon=ft.Icons.DELETE,
                    icon_color=ft.Colors.RED,
                    tooltip="Excluir exame",
                    on_click=lambda e, arquivo=exame['arquivo'], colaborador=exame['colaborador'],
                           exame_texto=exame['exame'], data=exame['data']:
                           self.excluir_exame(arquivo, colaborador, exame_texto, data)
                )
            ],
            alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
            vertical_alignment=ft.CrossAxisAlignment.CENTER,
            height=40
        )

    def build_content(self):
        main_content = ft.Column(
            [
                ft.Row(
                    [
                        ft.Text(
                            "Relações de Empresas",
                            size=36,
                            weight=ft.FontWeight.NORMAL,
                            color=ft.Colors.GREY_900,
                            font_family="Roboto"
                        )
                    ],
                    alignment=ft.MainAxisAlignment.CENTER,
                    vertical_alignment=ft.CrossAxisAlignment.CENTER,
                    height=80
                ),
                ft.Row(
                    [self.search_field],
                    alignment=ft.MainAxisAlignment.START,
                    vertical_alignment=ft.CrossAxisAlignment.CENTER,
                    height=60
                ),
                self.expansion_list
            ],
            spacing=10,
            scroll=ft.ScrollMode.AUTO
        )

        return ft.Container(
            content=ft.Row(
                [
                    ft.Container(
                        content=main_content,
                        bgcolor=ft.Colors.WHITE,          
                        border_radius=30,                  
                        padding=30,
                        margin=ft.margin.symmetric(horizontal=20, vertical=20),
                        expand=True,
                        alignment=ft.alignment.top_center
                    )
                ],
                width=self.page.width,
                height=self.page.height,
                alignment=ft.MainAxisAlignment.START
            ),
            bgcolor=ft.Colors.with_opacity(0.03, ft.Colors.GREY_100),
            expand=True
        )
