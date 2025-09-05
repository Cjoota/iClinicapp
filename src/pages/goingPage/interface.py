# ==============================================================================
# IMPORTAÇÕES DE BIBLIOTECAS
# ==============================================================================
import json
from dataclasses import dataclass
import base64
import shutil
import os
from datetime import datetime
from pathlib import Path

# Importações de bibliotecas de manipulação de imagem (Pillow)
from PIL import Image, ImageDraw, Image as PILImage

# Importações da biblioteca Flet para a interface gráfica
import flet as ft
import flet.canvas as cv

# Importações da biblioteca OpenPyXL para manipulação de ficheiros Excel
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import column_index_from_string
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.styles import Alignment
# Importações de outros módulos do projeto (assumindo a existência deles)
from src.functions.funcs import converter_xlsx_para_pdf, pdf_para_imagem_redimensionada
from src.utils.telaresize import Responsive

@dataclass
class State:
    """Uma classe simples para armazenar o estado (coordenadas x, y) do cursor no canvas."""
    x: float
    y: float

class Andamentos:
    """
    Controla a tela de "Andamentos", que exibe os exames pendentes do dia,
    permitindo a marcação de opções, captura de assinaturas e finalização dos exames.
    """

    def __init__(self, page: ft.Page):
        """
        Inicializa a classe Andamentos, configurando a página e carregando os dados iniciais.

        Args:
            page (ft.Page): A referência à página principal da aplicação Flet.
        """
        self.page = page

        # Carrega as configurações dos formulários a partir de ficheiros JSON
        self.json_anamnese = self.carregar_jsons("ANAMNESE")
        self.json_aso = self.carregar_jsons("ASO")
        self.json_audio = self.carregar_jsons("AUDIOMETRIA")
        
        # Estado inicial para o desenho no canvas (assinaturas, etc.)
        self.state = State(x=0,y=0)

        # Utilitários e dados iniciais
        self.responsive = Responsive(self.page)
        self.caminho_imagem_x = self.criar_imagem_x() # Gera e armazena o caminho para a imagem do 'X'
        self.empresas_exames = self.listar_empresas() # Carrega a lista de exames pendentes
        
        # --- Definição dos Elementos Visuais da Página --- 
        self.search_field = ft.TextField(
            hint_text="Pesquisar empresa",
            width=400,
            prefix_icon=ft.Icons.SEARCH,
            border_color=ft.Colors.GREY_400,
            bgcolor=ft.Colors.WHITE,
            color=ft.Colors.GREY_900,
            border_radius=8,
            height=48,
            on_change=self.filtrar_empresas
        )
        self.expansion_list = ft.ExpansionPanelList(
            expand_icon_color=ft.Colors.BLACK54,
            divider_color=ft.Colors.WHITE,
            elevation=10,
            on_change=self.expandir_empresas
        )
        
        # Carrega a lista inicial de empresas na tela
        self.carregar_empresa()

    def carregar_jsons(self, nome_arquivo: str) -> dict:
        """
        Carrega um ficheiro de configuração JSON a partir de um caminho predefinido.

        Args:
            nome_arquivo (str): O nome do ficheiro JSON (sem a extensão .json).

        Returns:
            dict: O conteúdo do ficheiro JSON como um dicionário, ou um dicionário vazio se o ficheiro não for encontrado.
        """
        caminho_json = Path(f"src/pages/goingPage/Jsons/{nome_arquivo}.json")
        if not caminho_json.exists():
            return {}
        with open(caminho_json, "r", encoding="utf-8") as f:
            return json.load(f)

    def criar_imagem_x(self, tamanho=15, cor="black", espessura=2, caminho_saida="assets/marcacao_x/x_mark.png"):
        """
        Cria uma imagem PNG de um 'X' se ela ainda não existir.
        Esta imagem é usada para marcar opções nos ficheiros Excel.

        Args:
            tamanho (int): A largura e altura da imagem em pixels.
            cor (str): A cor do 'X'.
            espessura (int): A espessura das linhas do 'X'.
            caminho_saida (str): O caminho onde a imagem será salva.

        Returns:
            str: O caminho para a imagem do 'X' gerada.
        """
        caminho_saida_path = Path(caminho_saida)
        if not caminho_saida_path.exists():
            caminho_saida_path.parent.mkdir(parents=True, exist_ok=True)
            img = Image.new("RGBA", (tamanho, tamanho), (255, 255, 255, 0))
            draw = ImageDraw.Draw(img)
            draw.line((0, 0, tamanho, tamanho), fill=cor, width=espessura)
            draw.line((0, tamanho, tamanho, 0), fill=cor, width=espessura)
            img.save(str(caminho_saida_path))
        return str(caminho_saida_path)
    
    def inserir_marcacao(self, ws, celula: str, caminho_imagem: str, offset_x_pixels=0, offset_y_pixels=0, v_align='top'):
        """
        Insere uma imagem numa célula. 
        Agora com um novo parâmetro 'v_align' para controlar o alinhamento vertical.
        """
        img = ExcelImage(caminho_imagem)
        col_letter = ''.join(filter(str.isalpha, celula))
        row_number = int(''.join(filter(str.isdigit, celula)))
        col = column_index_from_string(col_letter)

        offset_x_emu = int(offset_x_pixels * 9525)
        offset_y_emu = 0

        # Se pedirmos para alinhar em baixo, ele faz um cálculo especial
        if v_align == 'bottom':
            row_height_pt = ws.row_dimensions[row_number].height
            if row_height_pt is None:
                row_height_pt = 15 # Altura padrão do Excel

            cell_height_emu = row_height_pt * 12700
            image_height_emu = int(img.height * 9525)
            
            # Calcula o offset para que a imagem "encoste" na base
            offset_y_emu = cell_height_emu - image_height_emu
            offset_y_emu -= int(3 * 9525) # Pequena margem de 3 pixels
        
        # Se não, ele usa o alinhamento normal (pelo topo)
        else:
            offset_y_emu = int(offset_y_pixels * 9525)

        # O resto do código é o mesmo de antes
        marker = AnchorMarker(col=col-1, colOff=offset_x_emu, row=row_number-1, rowOff=offset_y_emu)
        size = XDRPositiveSize2D(cx=int(img.width * 9525), cy=int(img.height * 9525))
        anchor = OneCellAnchor(_from=marker, ext=size)
        img.anchor = anchor
        ws.add_image(img)

    def listar_empresas(self):
        """
        Lê a pasta 'documentos_gerados', filtra os exames do dia atual e os agrupa por empresa.

        Returns:
            dict: Um dicionário onde as chaves são nomes de empresas e os valores são listas de caminhos de ficheiros de exame.
        """
        
        # Lista as empresas que têm exames pendentes para o dia atual.
        documentos_dir = Path("documentos_gerados")
        empresas = {}
        if not documentos_dir.exists():
            return {}
        # Obtém a data atual para filtrar os exames do dia
        hoje = datetime.now().date()
        
        # Itera sobre todos os ficheiros .xlsx na pasta
        for arquivo in documentos_dir.glob("*.xlsx"):
            nome_arquivo = arquivo.stem
            partes = nome_arquivo.split()

            if len(partes) < 5:
                continue

            data_str = partes[-2]

            try:
                data_arquivo = datetime.strptime(data_str, "%d-%m-%Y").date()
            except ValueError:
                continue

            if data_arquivo == hoje:
                empresa = " ".join(partes[1:-3])
                if empresa not in empresas:
                    empresas[empresa] = []
                empresas[empresa].append(arquivo)
        
        # Retorna o dicionário de empresas com seus exames pendentes
        return empresas

    def criar_painel_empresa(self, empresa: str, cnpj):
        """Cria o componente visual (ExpansionPanel) para uma única empresa."""
        return ft.ExpansionPanel(
            can_tap_header=True,
            bgcolor="#EEFFEA",
            header=ft.Container(
                content=ft.Column(
                    [
                        ft.Text(
                            empresa.upper(),
                            size=18,
                            weight=ft.FontWeight.NORMAL,
                            color=ft.Colors.GREY_900,
                            font_family="Roboto"
                        ),
                        ft.Text(
                            cnpj,
                            size=14,
                            color=ft.Colors.GREY_600,
                            font_family="Roboto"
                        ) if cnpj else ft.Container()
                    ],
                    spacing=2,
                ),
                padding=ft.padding.symmetric(vertical=8, horizontal=12),
                bgcolor="#EEFFEA",
                border_radius=5
            ),
            content=ft.Container(
                content=ft.Text(
                    "Carregando documentos",
                    italic=True,
                    size=14,
                    color=ft.Colors.GREY_500
                ),
                bgcolor="#EEFFEA",
                padding=16
            )
        )

    def carregar_empresa(self, filtro=""):
        """Popula a lista principal de empresas na tela, aplicando um filtro de pesquisa."""
        self.expansion_list.controls.clear()
        for empresa, arquivos in self.empresas_exames.items():
            if filtro.lower() in empresa.lower():
                if len(self.expansion_list.controls) >= 50:
                    break
                painel = self.criar_painel_empresa(empresa, "")
                self.expansion_list.controls.append(painel)
        self.page.update()

    def filtrar_empresas(self, e):
        """É o gestor de eventos para o campo de pesquisa. Filtra a lista de empresas."""
        filtro = e.control.value
        self.carregar_empresa(filtro)

    def expandir_empresas(self, e):
        index = int(e.data)
        if index < 0: return

        painel_empresa = self.expansion_list.controls[index]
        if isinstance(painel_empresa.content, ft.ExpansionPanelList):
            self.page.update()
            return
            
        empresa = list(self.empresas_exames.keys())[index]
        exames_da_empresa = [ex for ex in self.listar_exames() if ex["empresa"] == empresa]

        exames_por_colaborador = {}
        for exame in exames_da_empresa:
            colaborador = exame['colaborador']
            if colaborador not in exames_por_colaborador:
                exames_por_colaborador[colaborador] = []
            exames_por_colaborador[colaborador].append(exame)

        paineis_colaboradores = []
        for colaborador, lista_de_exames in sorted(exames_por_colaborador.items()):
            # Cria o painel do colaborador primeiro
            painel_colaborador = ft.ExpansionPanel(
                can_tap_header=True,
                header=ft.ListTile(
                    leading=ft.Icon(ft.Icons.PERSON_OUTLINE),
                    title=ft.Text(colaborador, weight=ft.FontWeight.BOLD)
                )
            )

            linhas_de_exames_do_colaborador = []
            for exame in lista_de_exames:
                linha_exame = ft.Row(
                    alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                    vertical_alignment=ft.CrossAxisAlignment.CENTER,
                    spacing=10
                )
                # Passa o painel_colaborador para a função que cria os botões
                botoes = self.gerenciar_painel_exames(exame["arquivo"], exame["tipo_exame"], linha_exame, painel_colaborador)
                linha_exame.controls = [
                    ft.Text(f"Exame: {exame['tipo_exame']} | Data: {exame['data']}", size=14, color=ft.Colors.GREY_800, expand=True),
                    *botoes
                ]
                linhas_de_exames_do_colaborador.append(linha_exame)

            # Define o conteúdo do painel do colaborador
            painel_colaborador.content = ft.Container(
                content=ft.Column(linhas_de_exames_do_colaborador),
                padding=ft.padding.only(left=25, top=10, bottom=10, right=10)
            )
            paineis_colaboradores.append(painel_colaborador)

        painel_empresa.content = ft.ExpansionPanelList(
            controls=paineis_colaboradores,
            elevation=2
        )
        self.page.update()

    def listar_exames(self):
        """
        Lê a pasta 'documentos_gerados' e retorna uma lista detalhada de cada exame do dia.

        Returns:
            list: Uma lista de dicionários, onde cada dicionário representa um exame.
        """
        
        # Lista detalhada dos exames pendentes para o dia atual.
        documentos_dir = Path("documentos_gerados")
        exames = []
        
        # Verifica se a pasta existe
        if not documentos_dir.exists():
            return []
        
        # Obtém a data atual para filtrar os exames do dia
        hoje = datetime.now().date()
        
        # Itera sobre todos os ficheiros .xlsx na pasta
        for arquivo in documentos_dir.glob("*.xlsx"):
            nome_arquivo = arquivo.stem
            partes = nome_arquivo.split()

            if len(partes) < 5:
                continue

            data_str = partes[-2]

            try:
                data = datetime.strptime(data_str, "%d-%m-%Y").date()
            except ValueError:
                continue

            if data == hoje:
                tipo_exame = partes[0]
                colaborador = partes[-3]
                empresa = " ".join(partes[1:-3])

                exames.append({
                    "tipo_exame": tipo_exame,
                    "empresa": empresa,
                    "colaborador": colaborador,
                    "data": data.strftime("%d/%m/%Y"),
                    "arquivo": arquivo
                })
        return exames

    def abrir_dialog_assinatura(self, caminho_xlsx: Path, tipo_exame: str, quem_assina: str):
        """
        Abre um diálogo com um canvas para o utilizador desenhar a sua assinatura.
        """
        # Copia o arquivo para edição
        caminho_para_editar = self.preparar_arquivo_para_edicao(caminho_xlsx)
        
        # Cria o diálogo de assinatura
        linhas_desenhadas = []
        largura, altura = 400, 150  # tamanho do canvas para assinatura

        # Gestores de eventos para o desenho no canvas
        def pan_start(e: ft.DragStartEvent):
            self.state.x = e.local_x
            self.state.y = e.local_y
        
        async def pan_update(e: ft.DragUpdateEvent):
            # A função é assíncrona ('async') para garantir que a interface não trave durante o desenho rápido.
            linha = cv.Line(
                self.state.x, self.state.y, e.local_x, e.local_y,
                paint=ft.Paint(stroke_width=3, color=ft.Colors.BLACK)
            )
            linhas_desenhadas.append(linha)
            canvas.shapes.append(linha)
            canvas.update()
            self.state.x = e.local_x
            self.state.y = e.local_y

        def salvar_assinatura(e):
            """ Salva a assinatura e atualiza o ficheiro original. """
            try:
                # 1. Cria a imagem .png da assinatura (sem alterações)
                img = Image.new("RGBA", (largura, altura), (255, 255, 255, 0))
                draw = ImageDraw.Draw(img)
                for linha in linhas_desenhadas:
                    draw.line([(linha.x1, linha.y1), (linha.x2, linha.y2)], fill=(0, 0, 0, 255), width=3)
                
                nome_assinatura = f"{caminho_para_editar.stem}_assinatura_{quem_assina}.png"
                caminho_assinatura = Path("assets/assinaturas") / nome_assinatura
                caminho_assinatura.parent.mkdir(parents=True, exist_ok=True)
                img.save(str(caminho_assinatura))

                # 2. Insere a assinatura na CÓPIA do ficheiro Excel (sem alterações)
                self.inserir_assinatura_excel(caminho_para_editar, caminho_assinatura, tipo_exame, quem_assina)

                # --- 3. LINHA NOVA E CRUCIAL ---
                # Agora, movemos a cópia editada para substituir o ficheiro original.
                # Isto garante que o ficheiro principal está sempre atualizado.
                print(f"A atualizar o ficheiro original '{caminho_xlsx.name}' com a versão assinada.")
                shutil.move(str(caminho_para_editar), str(caminho_xlsx))
                # ---------------------------

                dialog_assinatura.open = False
                self.page.update()

            except Exception as ex:
                print(f"Erro ao salvar assinatura: {ex}")

        def limpar_canvas(e):
            """ Limpa todas as linhas desenhadas no canvas. """
            canvas.shapes.clear()
            linhas_desenhadas.clear()
            canvas.update()
        
        # Cria o canvas com o GestureDetector para capturar os eventos de desenho
        canvas = cv.Canvas(
            [],
            content=ft.GestureDetector(
                on_pan_start=pan_start,
                on_pan_update=pan_update,
                drag_interval=2,
                width=largura,
                height=altura,
            ),
            width=largura,
            height=altura,
        )
        
        # Container para o canvas com fundo branco
        canvas_container = ft.Container(
            content=canvas,
            width=largura,
            height=altura,
            bgcolor=ft.Colors.WHITE,
        )
        
        # Define o diálogo de assinatura
        dialog_assinatura = ft.AlertDialog(
            modal=True,
            title=ft.Text(f"Assinatura - {quem_assina.capitalize()}"),
            content=canvas_container,
            actions=[
                ft.TextButton("Limpar", on_click=limpar_canvas),
                ft.TextButton("Salvar", on_click=salvar_assinatura),
                ft.TextButton("Cancelar", on_click=lambda e: self.fechar_dialog(dialog_assinatura)),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        
        # Abre o diálogo
        self.page.open(dialog_assinatura)
        self.page.update()

    def inserir_assinatura_excel(self, caminho_xlsx: Path, caminho_assinatura: Path, tipo_exame: str, quem_assina: str):
        """ Insere a imagem da assinatura no ficheiro Excel na posição correta, dependendo do tipo de exame e quem assina. """
        try:
            wb = openpyxl.load_workbook(str(caminho_xlsx))
            # Seleciona a folha correta com base no tipo de exame
            if tipo_exame.upper() == "ANAMNESE":
                if len(wb.worksheets) > 1:
                    ws = wb.worksheets[1] 
                else:
                    ws = wb.worksheets[0]  
            else:
                ws = wb.worksheets[0]  

            # Redimensiona a imagem para caber na célula designada
            img_pil = PILImage.open(caminho_assinatura)
            largura_original, altura_original = img_pil.size

            # Define a largura desejada e calcula a altura proporcional
            largura_desejada = 150  
            altura_desejada = int(altura_original * (largura_desejada / largura_original))
            
            # Limita a altura máxima para evitar distorção excessiva
            img_excel = ExcelImage(str(caminho_assinatura))
            img_excel.width = largura_desejada
            img_excel.height = altura_desejada
            
            # Insere a imagem na célula correta com base no tipo de exame e quem assina
            if tipo_exame.upper() == "ANAMNESE":
                if quem_assina == "paciente":
                    ws.merge_cells('C51:D51')
                    ws.add_image(img_excel, 'C51')
            
            # Para outros tipos de exame, ajustar as células conforme necessário
            elif tipo_exame.upper() == "ASO":
                if quem_assina == "paciente":
                    ws.merge_cells('B58:C58')
                    ws.add_image(img_excel, 'B58')

            
            elif tipo_exame.upper() == "AUDIOMETRIA":
                if quem_assina == "paciente":
                    ws.merge_cells('O66:P66')
                    ws.add_image(img_excel, 'O66')

            else:
                print(f"Tipo de exame '{tipo_exame}' não reconhecido para inserir assinatura.")
                return

            wb.save(str(caminho_xlsx))
            print(f"Assinatura do {quem_assina} inserida no arquivo {caminho_xlsx} para o exame {tipo_exame}")
        
        except Exception as e:
            print(f"Erro ao inserir assinatura no Excel: {e}")

    def preparar_arquivo_para_edicao(self, caminho_original: Path) -> Path:
        
        """ Copia o arquivo original para uma pasta de edição, retornando o novo caminho. """
        
        # Cria a pasta de destino se não existir
        pasta_destino = Path("documentos_editados")
        pasta_destino.mkdir(parents=True, exist_ok=True)

        # Define o novo caminho para o arquivo copiado
        caminho_copia = pasta_destino / caminho_original.name

        # Copia o arquivo original para a pasta de edição
        shutil.copy2(caminho_original, caminho_copia)

        print(f"Arquivo copiado para edição: {caminho_copia}")

        # Retorna o caminho do arquivo copiado
        return caminho_copia

    def gerenciar_painel_exames(self, caminho_xlsx: Path, tipo_exame: str, linha_exame: ft.Row, painel_colaborador: ft.ExpansionPanel):
        """ Cria os botões de ação para cada exame, dependendo do tipo de exame. """

        botoes = []
        tipo = tipo_exame.upper()

        # Adiciona botões específicos com base no tipo de exame
        if tipo == "ANAMNESE":
            botoes.append(
                ft.ElevatedButton(
                    text="Marcar Opções ANAMNESE",
                    icon=ft.Icons.CHECK,
                    bgcolor="#2196F3",
                    color=ft.Colors.WHITE,
                    height=35,
                    on_click=lambda e, caminho=caminho_xlsx: self.controlar_anamnese(caminho)
                )
            )
            botoes.append(
                ft.ElevatedButton(
                    text="Assinatura Paciente",
                    icon=ft.Icons.BRUSH,
                    bgcolor="#4CAF50",
                    color=ft.Colors.BLACK,
                    height=35,
                    on_click=lambda e, caminho=caminho_xlsx, tipo=tipo: self.abrir_dialog_assinatura(caminho, tipo, "paciente")
                )
            )

        elif tipo == "ASO":
            botoes.append(
                ft.ElevatedButton(
                    text="Marcar opções ASO",
                    icon=ft.Icons.CHECK,
                    bgcolor="#2196F3",
                    color=ft.Colors.WHITE,
                    height=35,
                    on_click=lambda e, caminho=caminho_xlsx: self.controlar_aso(caminho)
                )
            )
            botoes.append(
                ft.ElevatedButton(
                    text="Assinatura Paciente",
                    icon=ft.Icons.BRUSH,
                    bgcolor="#4CAF50",
                    color=ft.Colors.BLACK,
                    height=35,
                    on_click=lambda e, caminho=caminho_xlsx, tipo=tipo: self.abrir_dialog_assinatura(caminho, tipo, "paciente")
                )
            )

        elif tipo == "AUDIOMETRIA":
            botoes.append(
                ft.ElevatedButton(
                    text="Marcar opções AUDIOMETRIA",
                    icon=ft.Icons.CHECK,
                    bgcolor="#2196F3",
                    color=ft.Colors.WHITE,
                    height=35,
                    on_click=lambda e, caminho=caminho_xlsx: self.controlar_audiometria(caminho)
                )
            )
            botoes.append(
                ft.ElevatedButton(
                    text="Assinatura Paciente",
                    icon=ft.Icons.BRUSH,
                    bgcolor="#4CAF50",
                    color=ft.Colors.BLACK,
                    height=35,
                    on_click=lambda e, caminho=caminho_xlsx, tipo=tipo: self.abrir_dialog_assinatura(caminho, tipo, "paciente")
                )
            )
        botoes.append(
            ft.ElevatedButton(
                text="Finalizar",
                icon=ft.Icons.DONE_ALL,
                bgcolor=ft.Colors.GREEN_700,
                color=ft.Colors.WHITE,
                # O on_click chama a nossa nova função, passando o ficheiro e a própria linha da UI
                on_click=lambda e, c=caminho_xlsx, r=linha_exame, p=painel_colaborador: self.finalizar_exame(c, r, p)
            )
        )   
        
        # Retorna a lista de botões criados
        return botoes
    def controlar_anamnese(self, caminho_xlsx: Path):
        caminho_x_img = self.caminho_imagem_x 
        lista_controles_geral, mapa_controles = [], {}
        
        for nome_pagina, questoes in self.json_anamnese.items():
            lista_controles_geral.append(ft.Text(nome_pagina, size=18, weight=ft.FontWeight.BOLD, color=ft.Colors.BLUE_GREY_700))
            for descricao, dados in questoes.items():
                tipo_controle = dados.get("tipo", "radio")
                controle = None
                
                if tipo_controle == "radio":
                    lista_controles_geral.append(ft.Text(descricao, size=16))
                    radios = [ft.Radio(value=cel, label=opt) for opt, cel in zip(dados["opcoes"], dados["celulas"])]
                    controle = ft.RadioGroup(content=ft.Row(radios))
                elif tipo_controle == "checkbox":
                    lista_controles_geral.append(ft.Text(descricao, size=16))
                    checkboxes = [ft.Checkbox(label=opt, data={"celula": cel}) for opt, cel in zip(dados["opcoes"], dados["celulas"])]
                    controle = ft.Row(controls=checkboxes)
                
                elif tipo_controle in ["texto", "texto_grande"]:
                    filtro = None
                    # Define o filtro com base no JSON
                    if dados.get("input_filter") == "digits": filtro = ft.InputFilter(allow=True, regex_string=r"[0-9]")
                    elif dados.get("input_filter") == "numbers": filtro = ft.InputFilter(allow=True, regex_string=r"[0-9.,]")
                    elif dados.get("input_filter") in ["pressure", "date"]: filtro = ft.InputFilter(allow=True, regex_string=r"[0-9/]")
                    
                    controle = ft.TextField(
                        label=descricao,
                        multiline=(tipo_controle == "texto_grande"),
                        min_lines=3 if tipo_controle == "texto_grande" else 1,
                        width=500 if tipo_controle == "texto_grande" else 300,
                        max_length=dados.get("max_length"),
                        input_filter=filtro,
                        suffix_text=dados.get("suffix"),
                        hint_text=dados.get("hint_text")
                    )
                
                if controle:
                    mapa_controles[descricao] = (controle, dados)
                    lista_controles_geral.append(controle)
                    lista_controles_geral.append(ft.Divider(height=10, color="transparent"))

        coluna_opcoes = ft.Column(lista_controles_geral, scroll=ft.ScrollMode.AUTO, height=500, spacing=5)

        def salvar_opcoes(e):
            # --- O SEU PAINEL DE CONTROLO COM AS SUAS MEDIDAS ---
            OFFSET_X_PADRAO = 6
            OFFSET_Y_PADRAO = 17
            offsets_excecoes = {
                "normal":    {"offset": (5, 20)},
                "escoliose": {"offset": (6, 30)},
                "lordose":   {"offset": (4, 6)},
                "cifose":    {"offset": (4, 100)},
                # Lembre-se de corrigir a chave abaixo para o NOME DA OPÇÃO da célula H43
                "nome_da_opcao_em_h43": {"offset": (5, 0), "v_align": "bottom"}
            }
            # ---------------------------------------------------

            def inserir_marcacao_com_excecao(ws, celula, label_opcao):
                config = offsets_excecoes.get(label_opcao.lower(), {})
                offset_x, offset_y = config.get("offset", (OFFSET_X_PADRAO, OFFSET_Y_PADRAO))
                v_align = config.get("v_align", "top")
                self.inserir_marcacao(ws, celula, caminho_x_img, 
                                      offset_x_pixels=offset_x, 
                                      offset_y_pixels=offset_y, 
                                      v_align=v_align)
            try:
                wb = openpyxl.load_workbook(str(caminho_xlsx))
                if len(wb.worksheets) > 1: ws = wb.worksheets[1]
                else: ws = wb.worksheets[0]

                for descricao, (controle, dados) in mapa_controles.items():
                    if isinstance(controle, ft.RadioGroup):
                        if controle.value:
                            self.inserir_marcacao(ws, controle.value, caminho_x_img, offset_x_pixels=OFFSET_X_PADRAO, offset_y_pixels=OFFSET_Y_PADRAO)
                    elif isinstance(controle, ft.Row) and all(isinstance(c, ft.Checkbox) for c in controle.controls):
                        for checkbox in controle.controls:
                            if checkbox.value:
                                inserir_marcacao_com_excecao(ws, checkbox.data["celula"], checkbox.label)
                    elif isinstance(controle, ft.TextField):
                        celula = dados["celula"]
                        valor_bruto = controle.value or ""
                        if valor_bruto:
                            sufixo = dados.get("suffix", "")
                            valor_final = f"{valor_bruto}{sufixo}"
                        else:
                            valor_final = ""
                        v_align = dados.get("vertical_align", "top")
                        cell_obj = ws[celula]
                        cell_obj.value = valor_final
                        cell_obj.alignment = Alignment(vertical=v_align, horizontal='center', wrap_text=True)

                wb.save(str(caminho_xlsx))
                dialog_opcoes.open = False
                self.page.update()
            except Exception as ex:
                print(f"Erro ao salvar opções ANAMNESE: {ex}")
        
        dialog_opcoes = ft.AlertDialog(
            modal=True,
            title=ft.Text("Marcar opções - ANAMNESE"),
            content=coluna_opcoes,
            actions=[
                ft.TextButton("Salvar", on_click=salvar_opcoes),
                ft.TextButton("Cancelar", on_click=lambda e: self.fechar_dialog(dialog_opcoes))
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )

        self.page.open(dialog_opcoes)
        self.page.update()
   
    def controlar_aso(self, caminho_xlsx: Path):
        """
        Abre o diálogo de formulário para o exame ASO, usando um RadioGroup para garantir
        que apenas uma opção possa ser selecionada.
        """
        caminho_x_img = self.criar_imagem_x()

        # --- CRIAÇÃO DA NOVA INTERFACE COM RADIOGROUP ---

        # 1. Prepara os controlos
        opcoes_radio = []
        # O campo de texto para a restrição é criado separadamente
        campo_texto_restricao = ft.TextField(
            label="Descreva a restrição",
            width=400,
            disabled=True # Começa desativado
        )
        # Vamos guardar a célula da opção "Apto com restrição" para referência
        celula_opcao_restricao = ""

        # 2. Itera sobre o JSON para criar os botões de rádio
        for descricao, dados in self.json_aso.items():
            # Caso especial para "Apto com restrição"
            if isinstance(dados, dict):
                celula_check = dados.get("check_cell")
                celula_texto = dados.get("text_cell")
                
                # Guarda a célula desta opção para a lógica do on_change
                celula_opcao_restricao = celula_check
                # O campo_texto_restricao já foi criado, só guardamos a sua célula de destino
                campo_texto_restricao.data = celula_texto
                
                # Cria o botão de rádio para esta opção
                radio = ft.Radio(label=descricao, value=celula_check)
                opcoes_radio.append(radio)

            # Para as outras opções simples
            else:
                celula = dados
                radio = ft.Radio(label=descricao, value=celula)
                opcoes_radio.append(radio)

        # 3. Função que é chamada QUANDO uma nova opção de rádio é selecionada
        def on_radio_change(e):
            # O valor do RadioGroup (e.control.value) é a célula da opção selecionada
            if e.control.value == celula_opcao_restricao:
                # Se a opção selecionada for a de restrição, ativa o campo de texto
                campo_texto_restricao.disabled = False
            else:
                # Se for qualquer outra, desativa e limpa o campo de texto
                campo_texto_restricao.disabled = True
                campo_texto_restricao.value = ""
            self.page.update()

        # 4. Cria o RadioGroup que vai conter todas as opções
        radio_group_aso = ft.RadioGroup(
            content=ft.Column(opcoes_radio),
            on_change=on_radio_change
        )
        
        # Monta a coluna final da UI
        coluna_opcoes = ft.Column(
            [
                radio_group_aso,
                campo_texto_restricao # Adiciona o campo de texto abaixo do grupo
            ], 
            scroll=ft.ScrollMode.AUTO, 
            height=400, 
            spacing=10
        )

        # --- LÓGICA DE SALVAR SIMPLIFICADA ---
        def salvar_opcoes(e):
            try:
                wb = openpyxl.load_workbook(str(caminho_xlsx))
                ws = wb.worksheets[0]

                # Pega o valor selecionado. Se nada for selecionado, não faz nada.
                celula_selecionada = radio_group_aso.value
                if not celula_selecionada:
                    # (Opcional) Pode mostrar um aviso aqui se quiser
                    print("Nenhuma opção selecionada.")
                    return

                # Insere a marcação 'X' na célula da opção escolhida
                self.inserir_marcacao(ws, celula_selecionada, caminho_x_img, offset_x_pixels=4, offset_y_pixels=3)

                # Se a opção escolhida foi a de restrição, salva o texto
                if celula_selecionada == celula_opcao_restricao:
                    celula_texto = campo_texto_restricao.data
                    texto_restricao = campo_texto_restricao.value or ""
                    ws[celula_texto] = texto_restricao
                
                wb.save(str(caminho_xlsx))
                dialog_opcoes.open = False
                self.page.update()
            except Exception as ex:
                print(f"Erro ao salvar opções ASO: {ex}")

        # O diálogo continua o mesmo
        dialog_opcoes = ft.AlertDialog(
            modal=True,
            title=ft.Text("Marcar opções - ASO"),
            content=coluna_opcoes,
            actions=[
                ft.TextButton("Salvar", on_click=salvar_opcoes),
                ft.TextButton("Cancelar", on_click=lambda e: self.fechar_dialog(dialog_opcoes)),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )

        self.page.open(dialog_opcoes)
        self.page.update()

    def controlar_audiometria(self, caminho_xlsx: str):
        """ Abre um diálogo para marcar opções de AUDIOMETRIA, lendo a configuração do JSON."""
        LARGURA_EXATA_PX = 1127
        ALTURA_EXATA_PX = 539
        
        # Caminho da imagem do 'X' para marcação
        caminho_x_img = self.caminho_imagem_x

        # Dicionário de offsets personalizados para cada célula
        offsets_personalizados = {
            "B54": (22, 8),
            "D56": (22, 4),
            "F56": (46, 3),
            "B57": (24, 15),
            "E57": (30, 15),
            "I57": (38, 15),
            "M57": (21, 15),
            "B60": (22, 8),
            "D62": (22, 4),
            "F62": (46, 3),
            "B63": (24, 11),
            "E63": (30, 11),
            "I63": (38, 11),
            "M63": (21, 11)
        }

        def inserir_marcacao_personalizada(ws, celula, caminho_imagem):
            """ Insere uma marcação com offsets personalizados para a audiometria. """
            offset_x, offset_y = offsets_personalizados.get(celula, (0, 0))
            self.inserir_marcacao(ws, celula, caminho_imagem, offset_x_pixels=offset_x, offset_y_pixels=offset_y)

        def salvar_desenho_audiometria(e, linhas_desenhadas, caminho_xlsx_alvo, dialog_para_fechar):
            """ Salva o desenho feito no canvas e insere a imagem no ficheiro Excel. """
            if not linhas_desenhadas:
                self.fechar_dialog(dialog_para_fechar)
                return

            try:
                caminho_imagem_fundo = Path("assets/audiometria/audiometria.png")
                pasta_saida_desenhos = Path("assets/desenhos_audiometria")
                pasta_saida_desenhos.mkdir(parents=True, exist_ok=True)
                nome_arquivo_saida = f"{Path(caminho_xlsx_alvo).stem}_desenho_final.png"
                caminho_imagem_saida = pasta_saida_desenhos / nome_arquivo_saida

                # Redimensiona a imagem de fundo para as dimensões exatas antes de compor
                fundo = Image.open(caminho_imagem_fundo).convert("RGBA")
                fundo = fundo.resize((LARGURA_EXATA_PX, ALTURA_EXATA_PX))

                # O overlay já é criado com o tamanho certo indiretamente pelo canvas
                desenho_overlay = Image.new("RGBA", (LARGURA_EXATA_PX, ALTURA_EXATA_PX), (255, 255, 255, 0))
                draw = ImageDraw.Draw(desenho_overlay)
                for linha in linhas_desenhadas:
                    draw.line([(linha.x1, linha.y1), (linha.x2, linha.y2)], fill=(255, 0, 0, 255), width=3)

                # Compõe a imagem final, que agora já tem o tamanho perfeito
                imagem_final = Image.alpha_composite(fundo, desenho_overlay)
                imagem_final.save(caminho_imagem_saida)

                # Insere no Excel (sem necessidade de forçar o tamanho, pois já está correto)
                wb = openpyxl.load_workbook(str(caminho_xlsx_alvo))
                ws = wb.worksheets[0]
                img_excel = ExcelImage(caminho_imagem_saida)

                # Opcional: A imagem já tem o tamanho certo, mas podemos garantir
                img_excel.width = LARGURA_EXATA_PX
                img_excel.height = ALTURA_EXATA_PX
                
                # Célula de destino confirmada
                celula_alvo = 'A22'
                ws.add_image(img_excel, celula_alvo)
                wb.save(str(caminho_xlsx_alvo))
                print(f"Imagem inserida com alinhamento perfeito em '{celula_alvo}'!")

            except Exception as ex:
                print(f"Erro ao salvar o desenho da audiometria: {ex}")
            finally:
                self.fechar_dialog(dialog_para_fechar)
        
        def fechar_dialog(dialog):
            """ Fecha o diálogo passado como parâmetro. """
            dialog.open = False
            self.page.update()

        def abrir_dialogo_tabela_audio(e=None):
            """ Abre o diálogo com o canvas para desenhar a audiometria. """
            
            # Lê e converte a imagem de fundo para base64 para uso no Flet
            caminho_imagem_path = Path("assets/audiometria/audiometria.png")
            with open(caminho_imagem_path, "rb") as f:
                imagem_bytes = f.read()
            imagem_b64 = base64.b64encode(imagem_bytes).decode("utf-8")

            # A lógica de desenho continua igual
            linhas_desenhadas = []
            last_x, last_y = 0, 0
            
            def pan_start(e: ft.DragStartEvent):
                """ Registra o ponto inicial do desenho. """
                nonlocal last_x, last_y
                last_x = e.local_x
                last_y = e.local_y
            
            async def pan_update(e: ft.DragUpdateEvent):
                """ A função é assíncrona ('async') para garantir que a interface não trave durante o desenho rápido. """
                nonlocal last_x, last_y
                linha = ft.canvas.Line(last_x, last_y, e.local_x, e.local_y, paint=ft.Paint(stroke_width=3, color=ft.Colors.RED))
                linhas_desenhadas.append(linha)
                canvas.shapes.append(linha)
                canvas.update()
                last_x, last_y = e.local_x, e.local_y
            
            def limpar_canvas(e):
                """ Limpa todas as linhas desenhadas no canvas. """
                canvas.shapes.clear()
                linhas_desenhadas.clear()
                canvas.update()

            # Cria o canvas e o GestureDetector para capturar os eventos de desenho
            canvas = ft.canvas.Canvas(
                [],
                width=LARGURA_EXATA_PX,
                height=ALTURA_EXATA_PX,
            )

            gesture_detector = ft.GestureDetector(
                content=canvas,
                on_pan_start=pan_start,
                on_pan_update=pan_update,
                drag_interval=2,
                width=LARGURA_EXATA_PX,
                height=ALTURA_EXATA_PX,
            )

            stack = ft.Stack(
                controls=[
                    ft.Image(
                        src_base64=imagem_b64,
                        width=LARGURA_EXATA_PX,
                        height=ALTURA_EXATA_PX,
                        # 'FILL' força a imagem a preencher a área, mesmo que estique.
                        # Isto garante que o seu guia visual corresponde perfeitamente ao canvas.
                        fit=ft.ImageFit.FILL,
                    ),
                    gesture_detector,
                ],
                width=LARGURA_EXATA_PX,
                height=ALTURA_EXATA_PX,
            )

            # Define o diálogo com o canvas
            dialog = ft.AlertDialog(
                modal=True,
                title=ft.Text("Tabela Audiometria - Marque as frequências"),
                content=stack,
                actions_alignment=ft.MainAxisAlignment.END,
            )
            
            dialog.actions = [
                ft.TextButton("Limpar", on_click=limpar_canvas),
                ft.TextButton("Fechar", on_click=lambda e: self.fechar_dialog(dialog)),
                ft.TextButton("Salvar", on_click=lambda e: salvar_desenho_audiometria(e, linhas_desenhadas, caminho_xlsx, dialog)),
            ]

            self.page.open(dialog)
            self.page.update()

        # Botão para abrir o diálogo da tabela audiometria
        botao_tabela_audio = ft.ElevatedButton("Tabela Audio", on_click=abrir_dialogo_tabela_audio)

        # Cria os controles de opções baseados no JSON
        controles_opcoes = []
        for secao, opcoes in self.json_audio.items():
            controles_opcoes.append(ft.Text(secao, style="headlineSmall", weight=ft.FontWeight.BOLD))
            grupo_secao = []
            for descricao, dados in opcoes.items():
                tipo = dados.get("tipo", "checkbox")
                celula = dados.get("celula")
                if tipo == "checkbox":
                    controle = ft.Checkbox(label=descricao, value=False)
                elif tipo == "texto":
                    controle = ft.TextField(label=descricao, width=400, multiline=True, max_lines=3)
                else:
                    controle = ft.Text(f"Tipo desconhecido para {descricao}")
                grupo_secao.append((descricao, celula, tipo, controle))
            controles_opcoes.append(grupo_secao)

        # Cria uma lista plana de controles para a coluna
        lista_controles = []
        for item in controles_opcoes:
            if isinstance(item, ft.Text):
                lista_controles.append(item)
            else:
                for _, _, _, controle in item:
                    lista_controles.append(controle)

        coluna_opcoes = ft.Column(lista_controles, scroll=ft.ScrollMode.AUTO, height=500)

        def salvar_opcoes(e):
            """ Salva as opções marcadas no ficheiro Excel. """
            
            try:
                wb = openpyxl.load_workbook(str(caminho_xlsx))
                ws = wb.worksheets[0]

                # Remove imagens antigas nas células de checkbox para evitar sobreposição
                todas_celulas_checkbox = []
                for item in controles_opcoes:
                    if isinstance(item, list):
                        for _, celula, tipo, _ in item:
                            if tipo == "checkbox":
                                todas_celulas_checkbox.append(celula)

                imagens_para_remover = []
                for img_excel in ws._images:
                    anchor = img_excel.anchor
                    if isinstance(anchor, str) and anchor in todas_celulas_checkbox:
                        imagens_para_remover.append(img_excel)
                    elif hasattr(anchor, "_from"):
                        col = anchor._from.col + 1
                        row = anchor._from.row + 1
                        cel = openpyxl.utils.get_column_letter(col) + str(row)
                        if cel in todas_celulas_checkbox:
                            imagens_para_remover.append(img_excel)
                for img_excel in imagens_para_remover:
                    ws._images.remove(img_excel)

                # Itera sobre os controles e insere as marcações no Excel
                for item in controles_opcoes:
                    if isinstance(item, list):
                        for descricao, celula, tipo, controle in item:
                            if tipo == "checkbox":
                                if controle.value:
                                    inserir_marcacao_personalizada(ws, celula, caminho_x_img)
                            elif tipo == "texto":
                                texto = controle.value or ""
                                ws[celula] = texto

                wb.save(str(caminho_xlsx))
                dialog_opcoes.open = False
                self.page.update()
            except Exception as ex:
                print(f"Erro ao salvar opções AUDIOMETRIA: {ex}")

        # Define o diálogo de opções
        dialog_opcoes = ft.AlertDialog(
            modal=True,
            title=ft.Text("Marcar opções - AUDIOMETRIA"),
            content=coluna_opcoes,
            actions=[
                ft.TextButton("Salvar", on_click=salvar_opcoes),
                ft.TextButton("Cancelar", on_click=lambda e: fechar_dialog(dialog_opcoes)),
                botao_tabela_audio  # botão para abrir o diálogo da tabela audiometria
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )

        self.page.open(dialog_opcoes)
        self.page.update()
        
    def finalizar_exame(self, caminho_xlsx: Path, linha_exame: ft.Row, painel_colaborador: ft.ExpansionPanel):
        """
        Move o ficheiro do exame e atualiza a UI, escondendo a linha do exame
        e também o painel do colaborador se ele ficar vazio.
        """
        print(f"--- 1. INICIANDO finalização para: {caminho_xlsx.name} ---")
        # A função interna _verificar_exame_assinado continua igual
        def _verificar_exame_assinado() -> bool:
            nome_assinatura = f"{caminho_xlsx.stem}_assinatura_paciente.png"
            caminho_assinatura = Path("assets/assinaturas") / nome_assinatura
            return caminho_assinatura.exists()

        if not _verificar_exame_assinado():
            # ... (lógica de aviso de assinatura em falta, sem alterações) ...
            return

        print("--- 2. Verificação de assinatura OK. A prosseguir para o bloco try... ---")
        try:
            # ... (lógica para mover o ficheiro e atualizar os dados, sem alterações) ...
            pasta_destino = Path("exames_prontos")
            pasta_destino.mkdir(parents=True, exist_ok=True)
            shutil.move(str(caminho_xlsx), str(pasta_destino))
            for empresa, arquivos in list(self.empresas_exames.items()):
                if caminho_xlsx in arquivos:
                    self.empresas_exames[empresa].remove(caminho_xlsx)
                    if not self.empresas_exames[empresa]:
                        del self.empresas_exames[empresa]
                    break
            print(f"--- 6. FICHEIRO MOVIDO para '{pasta_destino}' com sucesso. Atualizando UI... ---")
            # --- LÓGICA DE ATUALIZAÇÃO DA UI MELHORADA ---
            # 1. Esconde a linha do exame finalizado
            linha_exame.visible = False

            # 2. Verifica se ainda há exames visíveis para este colaborador
            # O caminho é: painel > content (Container) > content (Column) > controls (lista de Rows)
            exames_restantes = painel_colaborador.content.content.controls
            if all(not exame_row.visible for exame_row in exames_restantes):
                # Se todos estiverem invisíveis, esconde o painel do colaborador também
                painel_colaborador.visible = False

            self.page.snack_bar = ft.SnackBar(
                content=ft.Text(f"Exame '{caminho_xlsx.stem}' finalizado com sucesso!"),
                bgcolor=ft.Colors.GREEN_700,
            )
            self.page.snack_bar.open = True
            self.page.update()

        except Exception as ex:
            print(f"!!!!!!!! 7. ERRO INESPERADO OCORREU: {ex} !!!!!!!!")
            self.page.snack_bar = ft.SnackBar(
                content=ft.Text(f"Erro ao finalizar exame: {ex}"),
                bgcolor=ft.Colors.RED_700,
            )
            self.page.snack_bar.open = True
            self.page.update()

    def fechar_dialog(self, dialog):
        """ Fecha o diálogo passado como parâmetro. """
        
        dialog.open = False
        self.page.update()

    def build_content(self):
        """ Constrói o conteúdo principal da página. """
        
        # Cria o campo de pesquisa
        main_content = ft.Column(
            [
                ft.Row(
                    [
                        ft.Text(
                            "Andamentos",
                            size=36,
                            weight=ft.FontWeight.NORMAL,
                            color=ft.Colors.GREY_900,
                            font_family="Roboto"
                        )
                    ],
                    alignment=ft.MainAxisAlignment.CENTER,
                    vertical_alignment=ft.CrossAxisAlignment.CENTER,
                    height=80
                ),
                ft.Row(
                    [self.search_field],
                    alignment=ft.MainAxisAlignment.START,
                    vertical_alignment=ft.CrossAxisAlignment.CENTER,
                    height=60
                ),
                self.expansion_list
            ],
            spacing=10,
            scroll=ft.ScrollMode.AUTO
        )

        # Retorna o container principal com o conteúdo
        return ft.Container(
            content=ft.Row(
                [
                    ft.Container(
                        content=main_content,
                        bgcolor=ft.Colors.WHITE,
                        border_radius=30,
                        padding=30,
                        margin=ft.margin.symmetric(horizontal=20, vertical=20),
                        expand=True,
                        alignment=ft.alignment.top_center
                    )
                ],
                width=self.page.width,
                height=self.page.height,
                alignment=ft.MainAxisAlignment.START
            ),
            bgcolor=ft.Colors.with_opacity(0.03, ft.Colors.GREY_100),
            expand=True
        )