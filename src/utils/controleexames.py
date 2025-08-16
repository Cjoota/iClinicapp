import datetime
import re
from pathlib import Path
from openpyxl import load_workbook
import json
import shutil

class ControleExames:
    def __init__(self, pasta_export="exportados"):
        self.raiz = Path("relacoes")
        self.raiz_modelos = Path("relacoes/modelos")
        self.modelo = Path("relacoes/modelos/relacao.xlsx")
        self.json = Path("Relacao_empresas.json")
        self.pasta_export = Path(pasta_export)

        if not self.raiz.exists():
            self.raiz.mkdir(exist_ok=True)
        if not self.raiz_modelos.exists():
            self.raiz_modelos.mkdir(exist_ok=True)
        if not self.pasta_export.exists():
            self.pasta_export.mkdir(exist_ok=True)

        if datetime.datetime.now().day == 1:
            self.fechar_mes()
        self.init_json()

    def init_json(self):   
        dados = {} 
        if not self.json.exists():
            with open(self.json, "w", encoding="utf-8") as f:
                json.dump(dados, f, indent=4, ensure_ascii=False)

    def _carregar_json(self, empresa):
        try:
            with open(self.json, "r", encoding="utf-8") as f:
                dados = json.load(f)
                return dados[empresa]
        except:
            self._salvar_json(empresa, 6)
            with open(self.json, "r", encoding="utf-8") as f:
                dados = json.load(f)
                return dados[empresa]

    def _salvar_json(self, empresa, value):
        with open(self.json, "r", encoding="utf-8") as f:
            dados = json.load(f)
        dados[empresa] = value
        with open(self.json, "w", encoding="utf-8") as f:
            json.dump(dados, f, ensure_ascii=False, indent=4)

    def _criar_planilha(self, empresa: str):
        wb = load_workbook(self.modelo)
        ws = wb.active
        ws["B2"] = empresa
        empresa_arquivo = f"{empresa}_{datetime.datetime.now().strftime('%m-%Y')}.xlsx"
        saida = Path("relacoes")
        wb.save(saida / empresa_arquivo)

    def _limpar_texto_excel(self, texto):
        if not texto:
            return ""
        texto = re.sub(r'[~*/:?[\]\\<>|]', '', str(texto))
        return texto.strip()
    
    def registrar_exames(self, empresa: str, nome: str, exames: list[str], data_exame: str):
        relacao = Path(f"relacoes/{empresa}_{datetime.datetime.now().strftime('%m-%Y')}.xlsx")
        if not relacao.exists():
            self._criar_planilha(empresa)
            self._salvar_json(empresa, 6)
        
        wb = load_workbook(relacao)
        ws = wb.active
        inicial_value = self._carregar_json(empresa)
        
        ws[f"B{inicial_value}"] = self._limpar_texto_excel(nome)
        ws[f"D{inicial_value}"] = self._limpar_texto_excel(", ".join(exames) if isinstance(exames, list) else str(exames)) if not exames == [] else "ASO"
        ws[f"K{inicial_value}"] = self._limpar_texto_excel(data_exame)
        
        self._salvar_json(empresa, inicial_value + 1)

        wb.save(relacao)

    def fechar_mes(self):
        data_atual = datetime.datetime.now()
        if data_atual.day != 1:
            return  # só executa se for dia 1 do mês

        if not self.json.exists():
            return

        with open(self.json, "r", encoding="utf-8") as f:
            empresas_json = json.load(f)

        for empresa in empresas_json:
            nome_arquivo_atual = f"{empresa}_{data_atual.strftime('%m-%Y')}.xlsx"
            caminho_arquivo_atual = Path("relacoes") / nome_arquivo_atual

            if caminho_arquivo_atual.exists():
                destino = self.pasta_export / nome_arquivo_atual
                shutil.move(str(caminho_arquivo_atual), str(destino))

            self._criar_planilha(empresa)
            self._salvar_json(empresa, 6)  # resetar início na linha 6
